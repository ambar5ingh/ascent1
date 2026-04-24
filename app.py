"""
ASCENT — WRI India GHG Emissions & Scenario Planning Tool
IPCC 2019 / GPC Framework | AR6 GWP values
"""

from flask import Flask, jsonify, request, render_template, send_file
import io, json, math, os
import plotly.graph_objects as go
from plotly.subplots import make_subplots
import plotly
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

app = Flask(__name__)

# ─── AR6 GWP (from GWP sheet: CH4=29.8, N2O=273) ──────────────────────────────
GWP_CH4 = 29.8
GWP_N2O = 273.0

# ─── Emission Factors (from Emission Factor sheet) ────────────────────────────
# Format: { unique_id: (CO2_tTJ_or_tMWh, CH4_fraction, N2O_fraction, total_CO2e) }
# For electricity: t/MWh; for fuels: t/TJ
# CH4 and N2O stored as emission_rate * GWP already = tCO2e per TJ

EF = {
    # RESIDENTIAL
    "Res_Electricity":      {"co2": 0.823,  "ch4": 0.0,    "n2o": 0.0,    "unit": "t/MWh"},
    "Res_LPG":              {"co2": 63.1,   "ch4": 0.149,  "n2o": 0.0273, "unit": "t/TJ"},
    "Res_Firewood":         {"co2": 112.0,  "ch4": 8.94,   "n2o": 1.092,  "unit": "t/TJ"},
    "Res_Coal":             {"co2": 112.0,  "ch4": 5.96,   "n2o": 0.273,  "unit": "t/TJ"},
    "Res_PNG":              {"co2": 56.1,   "ch4": 0.149,  "n2o": 0.0273, "unit": "t/TJ"},
    "Res_Kerosene":         {"co2": 71.9,   "ch4": 0.298,  "n2o": 0.1638, "unit": "t/TJ"},
    "Res_Diesel_Genset":    {"co2": 74.1,   "ch4": 0.298,  "n2o": 0.1638, "unit": "t/TJ"},
    # COMMERCIAL
    "Com_Electricity":      {"co2": 0.823,  "ch4": 0.0,    "n2o": 0.0,    "unit": "t/MWh"},
    "Com_LPG":              {"co2": 63.1,   "ch4": 0.149,  "n2o": 0.0273, "unit": "t/TJ"},
    "Com_PNG":              {"co2": 56.1,   "ch4": 1.490,  "n2o": 0.0273, "unit": "t/TJ"},
    "Com_Firewood":         {"co2": 112.0,  "ch4": 8.94,   "n2o": 1.092,  "unit": "t/TJ"},
    "Com_Kerosene":         {"co2": 71.9,   "ch4": 0.298,  "n2o": 0.1638, "unit": "t/TJ"},
    # INSTITUTIONAL
    "Ins_Electricity":      {"co2": 0.823,  "ch4": 0.0,    "n2o": 0.0,    "unit": "t/MWh"},
    "Ins_LPG":              {"co2": 63.1,   "ch4": 0.149,  "n2o": 0.0273, "unit": "t/TJ"},
    # MANUFACTURING/INDUSTRIAL
    "Ind_Electricity":      {"co2": 0.823,  "ch4": 0.0,    "n2o": 0.0,    "unit": "t/MWh"},
    "Ind_LPG":              {"co2": 63.1,   "ch4": 0.0298, "n2o": 0.0273, "unit": "t/TJ"},
    "Ind_Coal":             {"co2": 112.0,  "ch4": 5.96,   "n2o": 1.092,  "unit": "t/TJ"},
    "Ind_Diesel":           {"co2": 74.1,   "ch4": 0.0894, "n2o": 0.1638, "unit": "t/TJ"},
    "Ind_PNG":              {"co2": 56.1,   "ch4": 0.0298, "n2o": 0.0273, "unit": "t/TJ"},
    "Ind_NatGas":           {"co2": 56.1,   "ch4": 0.0298, "n2o": 0.0273, "unit": "t/TJ"},
    # ENERGY GENERATION
    "EGen_Coal":            {"co2": 112.0,  "ch4": 5.96,   "n2o": 298.116,"unit": "t/TJ"},
    "EGen_NatGas":          {"co2": 56.1,   "ch4": 0.0298, "n2o": 7.4529, "unit": "t/TJ"},
    "EGen_Diesel":          {"co2": 74.1,   "ch4": 0.0894, "n2o": 44.717, "unit": "t/TJ"},
    "EGen_Paraffin":        {"co2": 73.3,   "ch4": 0.0894, "n2o": 44.717, "unit": "t/TJ"},
    "EGen_ResidualOil":     {"co2": 77.4,   "ch4": 0.0894, "n2o": 44.717, "unit": "t/TJ"},
    # TRANSPORT (from Base Year GHG Inventory sheet)
    "Trans_Petrol":         {"co2": 69.3,   "ch4": 0.0894, "n2o": 0.1638, "unit": "t/TJ", "conv_kl": 0.034839687},
    "Trans_Diesel":         {"co2": 74.1,   "ch4": 1.8476, "n2o": 0.0546, "unit": "t/TJ", "conv_kl": 0.038492544},
    "Trans_AutoLPG":        {"co2": 56.1,   "ch4": 8.94,   "n2o": 1.092,  "unit": "t/TJ", "conv_t":  0.0383791509},
    "Trans_CNG":            {"co2": 56.1,   "ch4": 5.96,   "n2o": 0.273,  "unit": "t/TJ", "conv_t":  0.048},
    "Trans_Electricity":    {"co2": 0.823,  "ch4": 0.0,    "n2o": 0.0,    "unit": "t/MWh"},
    "Trans_AvGasoline":     {"co2": 71.5,   "ch4": 0.0894, "n2o": 0.1638, "unit": "t/TJ", "conv_kl": 0.0334461},
    "Trans_JetKerosene":    {"co2": 71.5,   "ch4": 0.0894, "n2o": 0.1638, "unit": "t/TJ", "conv_kl": 0.37626862},
    # SOLID WASTE — Landfill CH4 method (IPCC 2006 FOD)
    # WASTEWATER — BOD/MCF method
}

# ─── Fuel conversion factors (from ListsAndTables, Cost and Conversion Factor) ─
FUEL_CONV = {
    # fuel_key: { "kl_to_tj": x, "t_to_tj": x } as applicable
    "Petrol":      {"kl_to_tj": 0.034839687},
    "Diesel":      {"kl_to_tj": 0.038492544},
    "AutoLPG":     {"t_to_tj":  0.0383791509},
    "CNG":         {"t_to_tj":  0.048},
    "LPG":         {"t_to_tj":  0.0473},
    "PNG":         {"t_to_tj":  0.038379151},   # town gas / city gas
    "Firewood":    {"t_to_tj":  0.020329320},
    "Coal":        {"t_to_tj":  0.029051584},
    "Kerosene":    {"kl_to_tj": 0.037630000},
    "NatGas":      {"t_to_tj":  0.048},
    "AvGasoline":  {"kl_to_tj": 0.0334461},
    "JetKerosene": {"kl_to_tj": 0.37626862},
}

# ─── Energy demand (kWh/m²/yr) by climate zone (from Assumption sheet) ────────
ENERGY_DEMAND = {
    "Hot and Dry": {
        "Residential":     {"fan":1,"pump":0,"cool":20,"heat":0,"equip":13,"light":7,"total":41},
        "Commercial":      {"fan":18,"pump":7,"cool":25,"heat":0,"equip":19,"light":8,"total":77},
        "Public/Inst":     {"fan":18,"pump":7,"cool":25,"heat":0,"equip":19,"light":8,"total":77},
        "Manufacturing":   {"fan":0,"pump":0,"cool":0,"heat":0,"equip":0,"light":0,"total":0},
    },
    "Warm and Humid": {
        "Residential":     {"fan":1,"pump":0,"cool":16,"heat":0,"equip":13,"light":7,"total":37},
        "Commercial":      {"fan":13,"pump":2,"cool":34,"heat":0,"equip":15,"light":6,"total":70},
        "Public/Inst":     {"fan":13,"pump":2,"cool":34,"heat":0,"equip":15,"light":6,"total":70},
        "Manufacturing":   {"fan":0,"pump":0,"cool":0,"heat":0,"equip":0,"light":0,"total":0},
    },
    "Composite": {
        "Residential":     {"fan":1,"pump":0,"cool":15,"heat":0,"equip":13,"light":7,"total":36},
        "Commercial":      {"fan":16,"pump":5,"cool":22,"heat":0,"equip":16,"light":7,"total":66},
        "Public/Inst":     {"fan":16,"pump":5,"cool":22,"heat":0,"equip":16,"light":7,"total":66},
        "Manufacturing":   {"fan":0,"pump":0,"cool":0,"heat":0,"equip":0,"light":0,"total":0},
    },
    "Temperate": {
        "Residential":     {"fan":0,"pump":0,"cool":8,"heat":0,"equip":13,"light":7,"total":28},
        "Commercial":      {"fan":17,"pump":5,"cool":18,"heat":0,"equip":18,"light":7,"total":65},
        "Public/Inst":     {"fan":17,"pump":5,"cool":18,"heat":0,"equip":18,"light":7,"total":65},
        "Manufacturing":   {"fan":0,"pump":0,"cool":0,"heat":0,"equip":0,"light":0,"total":0},
    },
    "Cold": {
        "Residential":     {"fan":0.8,"pump":0,"cool":6.1,"heat":12.3,"equip":13.4,"light":6.6,"total":39.2},
        "Commercial":      {"fan":15.2,"pump":4.5,"cool":9.6,"heat":17.6,"equip":12.7,"light":5.2,"total":64.8},
        "Public/Inst":     {"fan":15.2,"pump":4.5,"cool":9.6,"heat":17.6,"equip":12.7,"light":5.2,"total":64.8},
        "Manufacturing":   {"fan":0,"pump":0,"cool":0,"heat":0,"equip":0,"light":0,"total":0},
    },
}

# ─── Wastewater MCF values by treatment type (IPCC 2006 Table 6.3) ────────────
WW_MCF = {
    "aerobic_centralised": 0.0,
    "aerobic_ponds":       0.0,
    "anaerobic_lagoon":    0.8,
    "facultative_lagoon":  0.2,
    "constructed_wetland": 0.0,
    "anaerobic_reactor":   0.8,
    "sludge_anaerobic":    0.8,
    "composting":          0.0,
    "septic":              0.5,
    "open_pit":            0.1,
    "open_discharge":      0.06,  # rivers high organic
}

# ─── Solid Waste DOC values (IPCC 2006 Table 2.4) ─────────────────────────────
SW_DOC = {
    "food":    0.15,
    "garden":  0.20,
    "paper":   0.40,
    "wood":    0.43,
    "textile": 0.24,
    "rubber":  0.39,
}

# ─── IPPU emission factors ─────────────────────────────────────────────────────
IPPU_EF = {
    "cement_clinker":   0.510,   # tCO2/t clinker (IPCC 2019)
    "lime_high_ca":     0.785,
    "lime_dolomite":    0.913,
    "limestone":        0.480,   # tCO2/t
    "dolomite":         0.480,
    "steel_bof":        1.800,   # tCO2/t steel (IPCC default)
    "steel_eaf":        0.100,
    "ammonia":          1.694,   # tCO2/t (using IPCC fuel factor)
    "hno3_n2o":         9.0,     # kgN2O/t acid
    "glass_ef":         0.200,   # tCO2/t glass
}

# ─── AFOLU enteric fermentation EF (kg CH4/head/yr) — NATCOM II / IPCC 2006 ──
AFOLU_ENTERIC = {
    "dairy_cow_indigenous":   28,
    "nondairy_cow_adult":     32,
    "dairy_cow_crossbred":    43,
    "dairy_buffalo":          50,
    "sheep":                  5,
    "goat":                   5,
    "camel":                  46,
    "horse":                  18,
    "swine":                  1,
    "poultry":                0,
}
AFOLU_MANURE_CH4 = {
    "dairy_cow_indigenous":   3.5,
    "nondairy_cow_adult":     2.9,
    "dairy_cow_crossbred":    3.8,
    "dairy_buffalo":          4.4,
    "sheep":                  0.20,
    "goat":                   0.22,
    "camel":                  2.56,
    "horse":                  2.19,
    "swine":                  4.0,
    "poultry":                0.02,
}

# ─── Abatement cost assumptions (₹/tCO2e reduced, from Strategies & Cost) ────
ABATEMENT_COST = {
    "Buildings":  2500,   # ₹/tCO2e — weighted avg LED+appliance+solar
    "Transport":  3200,
    "Waste":      1800,
    "Wastewater": 1500,
    "AFOLU":      800,
    "IPPU":       4500,
}

# ─── India Cities Master (654 cities from City_master & expanded)
# State → District → City → Climate Zone
INDIA_CITIES = [
    {"state":"Andaman & Nicobar Islands","district":"South Andaman","city":"Port Blair","climate":"Warm & humid"},
    {"state":"Andhra Pradesh","district":"Anakapalli","city":"Narsipatnam","climate":"Warm & humid"},
    {"state":"Andhra Pradesh","district":"Anakapalli","city":"Yelamanchili","climate":"Warm & humid"},
    {"state":"Andhra Pradesh","district":"Anantapur","city":"Anantapur","climate":"Hot and Dry"},
    {"state":"Andhra Pradesh","district":"Anantapur","city":"Dharmavaram","climate":"Hot and Dry"},
    {"state":"Andhra Pradesh","district":"Anantapur","city":"Gooty","climate":"Hot and Dry"},
    {"state":"Andhra Pradesh","district":"Anantapur","city":"Guntakal","climate":"Hot and Dry"},
    {"state":"Andhra Pradesh","district":"Anantapur","city":"Kadiri","climate":"Hot and Dry"},
    {"state":"Andhra Pradesh","district":"Anantapur","city":"Kalyandurgam","climate":"Hot and Dry"},
    {"state":"Andhra Pradesh","district":"Anantapur","city":"Rayadurg","climate":"Hot and Dry"},
    {"state":"Andhra Pradesh","district":"Anantapur","city":"Tadipatri","climate":"Hot and Dry"},
    {"state":"Andhra Pradesh","district":"Annamaiya","city":"B Kothakota","climate":"Warm & humid"},
    {"state":"Andhra Pradesh","district":"Annamaiya","city":"Madanapalle","climate":"Warm & humid"},
    {"state":"Andhra Pradesh","district":"Annamaiya","city":"Rajampeta","climate":"Warm & humid"},
    {"state":"Andhra Pradesh","district":"Annamaiya","city":"Rayachoty","climate":"Warm & humid"},
    {"state":"Andhra Pradesh","district":"Bapatla","city":"Addanki","climate":"Warm & humid"},
    {"state":"Andhra Pradesh","district":"Bapatla","city":"Repalle","climate":"Warm & humid"},
    {"state":"Andhra Pradesh","district":"Chittoor","city":"Chittoor","climate":"Warm & humid"},
    {"state":"Andhra Pradesh","district":"Chittoor","city":"Kuppam","climate":"Warm & humid"},
    {"state":"Andhra Pradesh","district":"Chittoor","city":"Nagari","climate":"Warm & humid"},
    {"state":"Andhra Pradesh","district":"Chittoor","city":"Palamaner","climate":"Warm & humid"},
    {"state":"Andhra Pradesh","district":"Chittoor","city":"Punganur","climate":"Warm & humid"},
    {"state":"Andhra Pradesh","district":"East Godavari","city":"Kovvur","climate":"Warm & humid"},
    {"state":"Andhra Pradesh","district":"East Godavari","city":"Nidadavole","climate":"Warm & humid"},
    {"state":"Andhra Pradesh","district":"East Godavari","city":"Rajahmundry","climate":"Warm & humid"},
    {"state":"Andhra Pradesh","district":"Eluru","city":"Chintalapudi","climate":"Warm & humid"},
    {"state":"Andhra Pradesh","district":"Eluru","city":"Eluru","climate":"Warm & humid"},
    {"state":"Andhra Pradesh","district":"Eluru","city":"Jangareddygudem","climate":"Warm & humid"},
    {"state":"Andhra Pradesh","district":"Eluru","city":"Nuzivid","climate":"Warm & humid"},
    {"state":"Andhra Pradesh","district":"Guntur","city":"Bapatla","climate":"Warm & humid"},
    {"state":"Andhra Pradesh","district":"Guntur","city":"Guntur","climate":"Warm & humid"},
    {"state":"Andhra Pradesh","district":"Guntur","city":"Mangalagiri Tadepalli","climate":"Warm & humid"},
    {"state":"Andhra Pradesh","district":"Guntur","city":"Ponnur","climate":"Warm & humid"},
    {"state":"Andhra Pradesh","district":"Guntur","city":"Tenali","climate":"Warm & humid"},
    {"state":"Andhra Pradesh","district":"Kadapa","city":"Badvel","climate":"Warm & humid"},
    {"state":"Andhra Pradesh","district":"Kadapa","city":"Jammalamadugu","climate":"Warm & humid"},
    {"state":"Andhra Pradesh","district":"Kadapa","city":"Kadapa","climate":"Warm & humid"},
    {"state":"Andhra Pradesh","district":"Kadapa","city":"Kamalapuram","climate":"Warm & humid"},
    {"state":"Andhra Pradesh","district":"Kadapa","city":"Mydukur","climate":"Warm & humid"},
    {"state":"Andhra Pradesh","district":"Kadapa","city":"Proddatur","climate":"Warm & humid"},
    {"state":"Andhra Pradesh","district":"Kadapa","city":"Pulivendula","climate":"Warm & humid"},
    {"state":"Andhra Pradesh","district":"Kakinada","city":"Kakinada","climate":"Warm & humid"},
    {"state":"Andhra Pradesh","district":"Kakinada","city":"Pithapuram","climate":"Warm & humid"},
    {"state":"Andhra Pradesh","district":"Kakinada","city":"Samalkot","climate":"Warm & humid"},
    {"state":"Andhra Pradesh","district":"Krishna","city":"Gudivada","climate":"Warm & humid"},
    {"state":"Andhra Pradesh","district":"Krishna","city":"Machilipatnam","climate":"Warm & humid"},
    {"state":"Andhra Pradesh","district":"Krishna","city":"Vijayawada","climate":"Warm & humid"},
    {"state":"Andhra Pradesh","district":"Kurnool","city":"Adoni","climate":"Warm & humid"},
    {"state":"Andhra Pradesh","district":"Kurnool","city":"Kurnool","climate":"Warm & humid"},
    {"state":"Andhra Pradesh","district":"NTR","city":"Vijayawada","climate":"Warm & humid"},
    {"state":"Andhra Pradesh","district":"Nellore","city":"Nellore","climate":"Warm & humid"},
    {"state":"Andhra Pradesh","district":"Prakasam","city":"Ongole","climate":"Warm & humid"},
    {"state":"Andhra Pradesh","district":"Srikakulam","city":"Srikakulam","climate":"Warm & humid"},
    {"state":"Andhra Pradesh","district":"Tirupati","city":"Tirupati","climate":"Warm & humid"},
    {"state":"Andhra Pradesh","district":"Visakhapatnam","city":"Gvmc Visakhapatnam","climate":"Warm & humid"},
    {"state":"Andhra Pradesh","district":"Vizianagaram","city":"Vizianagaram","climate":"Warm & humid"},
    {"state":"Andhra Pradesh","district":"West Godavari","city":"Bhimavaram","climate":"Warm & humid"},
    {"state":"Andhra Pradesh","district":"West Godavari","city":"Tadepalligudem","climate":"Warm & humid"},
    {"state":"Arunachal Pradesh","district":"East Siang","city":"Pasighat","climate":"Cold"},
    {"state":"Arunachal Pradesh","district":"Papum Pare","city":"Itanagar","climate":"Cold"},
    {"state":"Arunachal Pradesh","district":"Tawang","city":"Tawang","climate":"Cold"},
    {"state":"Arunachal Pradesh","district":"West Kameng","city":"Bomdila","climate":"Cold"},
    {"state":"Assam","district":"Barpeta","city":"Barpeta","climate":"Warm & humid"},
    {"state":"Assam","district":"Barpeta","city":"Barpeta Road","climate":"Warm & humid"},
    {"state":"Assam","district":"Bongaigaon","city":"Bongaigaon","climate":"Warm & humid"},
    {"state":"Assam","district":"Cachar","city":"Silchar","climate":"Warm & humid"},
    {"state":"Assam","district":"Dibrugarh","city":"Dibrugarh","climate":"Warm & humid"},
    {"state":"Assam","district":"Golaghat","city":"Golaghat","climate":"Warm & humid"},
    {"state":"Assam","district":"Jorhat","city":"Jorhat","climate":"Warm & humid"},
    {"state":"Assam","district":"Kamrup","city":"Guwahati","climate":"Warm & humid"},
    {"state":"Assam","district":"Kamrup","city":"North Guwahati","climate":"Warm & humid"},
    {"state":"Assam","district":"Karbi Anlong","city":"Diphu","climate":"Warm & humid"},
    {"state":"Assam","district":"Karimganj","city":"Karimganj","climate":"Warm & humid"},
    {"state":"Assam","district":"Kokrajhar","city":"Kokrajhar","climate":"Warm & humid"},
    {"state":"Assam","district":"Nagaon","city":"Nagaon","climate":"Warm & humid"},
    {"state":"Assam","district":"Nalbari","city":"Nalbari","climate":"Warm & humid"},
    {"state":"Assam","district":"North Lakhimpur","city":"North Lakhimpur","climate":"Warm & humid"},
    {"state":"Assam","district":"Sibsagar","city":"Sibsagar","climate":"Warm & humid"},
    {"state":"Assam","district":"Sonitpur","city":"Tezpur","climate":"Warm & humid"},
    {"state":"Assam","district":"Tinsukia","city":"Tinsukia","climate":"Warm & humid"},
    {"state":"Bihar","district":"Araria","city":"Araria","climate":"Composite"},
    {"state":"Bihar","district":"Aurangabad","city":"Aurangabad","climate":"Composite"},
    {"state":"Bihar","district":"Begusarai","city":"Begusarai","climate":"Composite"},
    {"state":"Bihar","district":"Bhagalpur","city":"Bhagalpur","climate":"Composite"},
    {"state":"Bihar","district":"Bhojpur","city":"Ara","climate":"Composite"},
    {"state":"Bihar","district":"Buxar","city":"Buxar","climate":"Composite"},
    {"state":"Bihar","district":"Darbhanga","city":"Darbhanga","climate":"Composite"},
    {"state":"Bihar","district":"Gaya","city":"Bodh Gaya","climate":"Composite"},
    {"state":"Bihar","district":"Gaya","city":"Gaya","climate":"Composite"},
    {"state":"Bihar","district":"Gopalganj","city":"Gopalganj","climate":"Composite"},
    {"state":"Bihar","district":"Jahanabad","city":"Jehanabad","climate":"Composite"},
    {"state":"Bihar","district":"Jamui","city":"Jamui","climate":"Composite"},
    {"state":"Bihar","district":"Katihar","city":"Katihar","climate":"Composite"},
    {"state":"Bihar","district":"Khagaria","city":"Khagaria","climate":"Composite"},
    {"state":"Bihar","district":"Kishanganj","city":"Kishanganj","climate":"Composite"},
    {"state":"Bihar","district":"Lakhisarai","city":"Lakhisarai","climate":"Composite"},
    {"state":"Bihar","district":"Madhepura","city":"Madhepura","climate":"Composite"},
    {"state":"Bihar","district":"Madhubani","city":"Madhubani","climate":"Composite"},
    {"state":"Bihar","district":"Munger","city":"Munger","climate":"Composite"},
    {"state":"Bihar","district":"Muzaffarpur","city":"Muzaffarpur","climate":"Composite"},
    {"state":"Bihar","district":"Nalanda","city":"Biharsharif","climate":"Composite"},
    {"state":"Bihar","district":"Nalanda","city":"Rajgir","climate":"Composite"},
    {"state":"Bihar","district":"Nawada","city":"Nawada","climate":"Composite"},
    {"state":"Bihar","district":"Paschim Champaran","city":"Bettiah","climate":"Composite"},
    {"state":"Bihar","district":"Patna","city":"Patna","climate":"Composite"},
    {"state":"Bihar","district":"Purnia","city":"Purnia","climate":"Composite"},
    {"state":"Bihar","district":"Purvi Champaran","city":"Motihari","climate":"Composite"},
    {"state":"Bihar","district":"Rohtas","city":"Sasaram","climate":"Composite"},
    {"state":"Bihar","district":"Saharsa","city":"Saharsa","climate":"Composite"},
    {"state":"Bihar","district":"Samastipur","city":"Samastipur","climate":"Composite"},
    {"state":"Bihar","district":"Saran","city":"Chapra","climate":"Composite"},
    {"state":"Bihar","district":"Sheohar","city":"Sheohar","climate":"Composite"},
    {"state":"Bihar","district":"Sitamarhi","city":"Sitamarhi","climate":"Composite"},
    {"state":"Bihar","district":"Siwan","city":"Siwan","climate":"Composite"},
    {"state":"Bihar","district":"Supaul","city":"Supaul","climate":"Composite"},
    {"state":"Bihar","district":"Vaishali","city":"Hajipur","climate":"Composite"},
    {"state":"Chandigarh","district":"Chandigarh","city":"Chandigarh","climate":"Composite"},
    {"state":"Chhattisgarh","district":"Balod","city":"Balod","climate":"Composite"},
    {"state":"Chhattisgarh","district":"Bilaspur","city":"Bilaspur","climate":"Composite"},
    {"state":"Chhattisgarh","district":"Durg","city":"Bhilai Nagar","climate":"Composite"},
    {"state":"Chhattisgarh","district":"Durg","city":"Durg","climate":"Composite"},
    {"state":"Chhattisgarh","district":"Jagdalpur","city":"Jagdalpur","climate":"Composite"},
    {"state":"Chhattisgarh","district":"Janjgir Champa","city":"Champa","climate":"Composite"},
    {"state":"Chhattisgarh","district":"Korba","city":"Korba","climate":"Composite"},
    {"state":"Chhattisgarh","district":"Raigarh","city":"Raigarh","climate":"Composite"},
    {"state":"Chhattisgarh","district":"Raipur","city":"Raipur","climate":"Composite"},
    {"state":"Chhattisgarh","district":"Rajnandgaon","city":"Rajnandgaon","climate":"Composite"},
    {"state":"Chhattisgarh","district":"Sarguja","city":"Ambikapur","climate":"Composite"},
    {"state":"Dadra & Nagar Haveli","district":"Dadra And Nagar Haveli","city":"Silvassa","climate":"Warm & humid"},
    {"state":"Daman & Diu","district":"Daman","city":"Daman","climate":"Warm & humid"},
    {"state":"Daman & Diu","district":"Diu","city":"Diu","climate":"Warm & humid"},
    {"state":"Delhi","district":"New Delhi","city":"New Delhi","climate":"Composite"},
    {"state":"Delhi","district":"South Delhi","city":"Municipal Corporation Of Delhi","climate":"Composite"},
    {"state":"Goa","district":"North Goa","city":"Mapusa","climate":"Warm & humid"},
    {"state":"Goa","district":"North Goa","city":"Panaji","climate":"Warm & humid"},
    {"state":"Goa","district":"South Goa","city":"Margao","climate":"Warm & humid"},
    {"state":"Goa","district":"South Goa","city":"Mormugao","climate":"Warm & humid"},
    {"state":"Gujarat","district":"Ahmedabad","city":"Ahmedabad","climate":"Hot and Dry"},
    {"state":"Gujarat","district":"Amreli","city":"Amreli","climate":"Hot and Dry"},
    {"state":"Gujarat","district":"Anand","city":"Anand","climate":"Hot and Dry"},
    {"state":"Gujarat","district":"Anand","city":"Vallabh Vidhyanagar","climate":"Hot and Dry"},
    {"state":"Gujarat","district":"Aravalli","city":"Modasa","climate":"Hot and Dry"},
    {"state":"Gujarat","district":"Banas Kantha","city":"Palanpur","climate":"Hot and Dry"},
    {"state":"Gujarat","district":"Bharuch","city":"Ankleshwer","climate":"Hot and Dry"},
    {"state":"Gujarat","district":"Bharuch","city":"Bharuch","climate":"Hot and Dry"},
    {"state":"Gujarat","district":"Bhavnagar","city":"Bhavnagar","climate":"Hot and Dry"},
    {"state":"Gujarat","district":"Botad","city":"Botad","climate":"Hot and Dry"},
    {"state":"Gujarat","district":"Chhota Udepure","city":"Chhota Udepur","climate":"Hot and Dry"},
    {"state":"Gujarat","district":"Dahod","city":"Dahod","climate":"Hot and Dry"},
    {"state":"Gujarat","district":"Devbhoomi Dwarka","city":"Dwarka","climate":"Hot and Dry"},
    {"state":"Gujarat","district":"Gandhinagar","city":"Gandhinagar","climate":"Hot and Dry"},
    {"state":"Gujarat","district":"Gandhinagar","city":"Kalol","climate":"Hot and Dry"},
    {"state":"Gujarat","district":"Gir Somnath","city":"Veraval","climate":"Hot and Dry"},
    {"state":"Gujarat","district":"Jamnagar","city":"Jamnagar","climate":"Hot and Dry"},
    {"state":"Gujarat","district":"Junagadh","city":"Junagadh","climate":"Hot and Dry"},
    {"state":"Gujarat","district":"Kachchh","city":"Bhuj","climate":"Hot and Dry"},
    {"state":"Gujarat","district":"Kachchh","city":"Gandhidham","climate":"Hot and Dry"},
    {"state":"Gujarat","district":"Kheda","city":"Nadiad","climate":"Hot and Dry"},
    {"state":"Gujarat","district":"Mahesana","city":"Mahesana","climate":"Hot and Dry"},
    {"state":"Gujarat","district":"Morbi","city":"Morbi","climate":"Hot and Dry"},
    {"state":"Gujarat","district":"Narmada","city":"Rajpipla","climate":"Hot and Dry"},
    {"state":"Gujarat","district":"Navsari","city":"Navsari","climate":"Hot and Dry"},
    {"state":"Gujarat","district":"Panch Mahals","city":"Godhra","climate":"Hot and Dry"},
    {"state":"Gujarat","district":"Patan","city":"Patan","climate":"Hot and Dry"},
    {"state":"Gujarat","district":"Porbandar","city":"Porbandar","climate":"Hot and Dry"},
    {"state":"Gujarat","district":"Rajkot","city":"Gondal","climate":"Hot and Dry"},
    {"state":"Gujarat","district":"Rajkot","city":"Rajkot","climate":"Hot and Dry"},
    {"state":"Gujarat","district":"Sabar Kantha","city":"Himmatnagar","climate":"Hot and Dry"},
    {"state":"Gujarat","district":"Surat","city":"Bardoli","climate":"Hot and Dry"},
    {"state":"Gujarat","district":"Surat","city":"Surat","climate":"Hot and Dry"},
    {"state":"Gujarat","district":"Surendranagar Dudhrej","city":"Surendranagar","climate":"Hot and Dry"},
    {"state":"Gujarat","district":"Vadodara","city":"Vadodara","climate":"Hot and Dry"},
    {"state":"Gujarat","district":"Valsad","city":"Valsad","climate":"Hot and Dry"},
    {"state":"Gujarat","district":"Valsad","city":"Vapi","climate":"Hot and Dry"},
    {"state":"Haryana","district":"Ambala","city":"Ambala","climate":"Composite"},
    {"state":"Haryana","district":"Bhiwani","city":"Bhiwani","climate":"Composite"},
    {"state":"Haryana","district":"Faridabad","city":"Faridabad","climate":"Composite"},
    {"state":"Haryana","district":"Gurugram","city":"Gurugram","climate":"Composite"},
    {"state":"Haryana","district":"Hisar","city":"Hisar","climate":"Composite"},
    {"state":"Haryana","district":"Jhajjar","city":"Bahadurgarh","climate":"Composite"},
    {"state":"Haryana","district":"Jind","city":"Jind","climate":"Composite"},
    {"state":"Haryana","district":"Kaithal","city":"Kaithal","climate":"Composite"},
    {"state":"Haryana","district":"Karnal","city":"Karnal","climate":"Composite"},
    {"state":"Haryana","district":"Kurukshetra","city":"Thanesar","climate":"Composite"},
    {"state":"Haryana","district":"Mahendragarh","city":"Narnaul","climate":"Composite"},
    {"state":"Haryana","district":"Nuh","city":"Nuh","climate":"Composite"},
    {"state":"Haryana","district":"Palwal","city":"Palwal","climate":"Composite"},
    {"state":"Haryana","district":"Panchkula","city":"Panchkula","climate":"Composite"},
    {"state":"Haryana","district":"Panipat","city":"Panipat","climate":"Composite"},
    {"state":"Haryana","district":"Rewari","city":"Rewari","climate":"Composite"},
    {"state":"Haryana","district":"Rohtak","city":"Rohtak","climate":"Composite"},
    {"state":"Haryana","district":"Sirsa","city":"Sirsa","climate":"Composite"},
    {"state":"Haryana","district":"Sonipat","city":"Sonipat","climate":"Composite"},
    {"state":"Haryana","district":"Yamunanagar","city":"Yamunanagar","climate":"Composite"},
    {"state":"Himachal Pradesh","district":"Bilaspur","city":"Bilaspur","climate":"Cold"},
    {"state":"Himachal Pradesh","district":"Chamba","city":"Chamba","climate":"Cold"},
    {"state":"Himachal Pradesh","district":"Hamirpur","city":"Hamirpur","climate":"Cold"},
    {"state":"Himachal Pradesh","district":"Kangra","city":"Dharmsala","climate":"Temperate"},
    {"state":"Himachal Pradesh","district":"Kangra","city":"Kangra","climate":"Temperate"},
    {"state":"Himachal Pradesh","district":"Kullu","city":"Kullu","climate":"Cold"},
    {"state":"Himachal Pradesh","district":"Kullu","city":"Manali","climate":"Cold"},
    {"state":"Himachal Pradesh","district":"Mandi","city":"Mandi","climate":"Temperate"},
    {"state":"Himachal Pradesh","district":"Shimla","city":"Shimla","climate":"Cold"},
    {"state":"Himachal Pradesh","district":"Sirmour","city":"Nahan","climate":"Cold"},
    {"state":"Himachal Pradesh","district":"Solan","city":"Solan","climate":"Cold"},
    {"state":"Himachal Pradesh","district":"Una","city":"Una","climate":"Cold"},
    {"state":"Jammu & Kashmir","district":"Anantnag","city":"Anantnag","climate":"Cold"},
    {"state":"Jammu & Kashmir","district":"Badgam","city":"Badgam","climate":"Cold"},
    {"state":"Jammu & Kashmir","district":"Baramula","city":"Baramula","climate":"Cold"},
    {"state":"Jammu & Kashmir","district":"Baramula","city":"Sopore","climate":"Cold"},
    {"state":"Jammu & Kashmir","district":"Doda","city":"Doda","climate":"Cold"},
    {"state":"Jammu & Kashmir","district":"Jammu","city":"Jammu","climate":"Cold"},
    {"state":"Jammu & Kashmir","district":"Kathua","city":"Kathua","climate":"Cold"},
    {"state":"Jammu & Kashmir","district":"Kulgam","city":"Kulgam","climate":"Cold"},
    {"state":"Jammu & Kashmir","district":"Kupwara","city":"Kupwara","climate":"Cold"},
    {"state":"Jammu & Kashmir","district":"Pulwama","city":"Pulwama","climate":"Cold"},
    {"state":"Jammu & Kashmir","district":"Punch","city":"Punch","climate":"Cold"},
    {"state":"Jammu & Kashmir","district":"Rajouri","city":"Rajouri","climate":"Cold"},
    {"state":"Jammu & Kashmir","district":"Srinagar","city":"Srinagar","climate":"Cold"},
    {"state":"Jammu & Kashmir","district":"Udhampur","city":"Udhampur","climate":"Cold"},
    {"state":"Jharkhand","district":"Bokaro","city":"Chas","climate":"Composite"},
    {"state":"Jharkhand","district":"Chatra","city":"Chatra","climate":"Composite"},
    {"state":"Jharkhand","district":"Deoghar","city":"Deoghar","climate":"Composite"},
    {"state":"Jharkhand","district":"Dhanbad","city":"Dhanbad","climate":"Composite"},
    {"state":"Jharkhand","district":"Dumka","city":"Dumka","climate":"Composite"},
    {"state":"Jharkhand","district":"East Singhbhum","city":"Jamshedpur","climate":"Composite"},
    {"state":"Jharkhand","district":"Garhwa","city":"Garhwa","climate":"Composite"},
    {"state":"Jharkhand","district":"Giridih","city":"Giridih","climate":"Composite"},
    {"state":"Jharkhand","district":"Godda","city":"Godda","climate":"Composite"},
    {"state":"Jharkhand","district":"Gumla","city":"Gumla","climate":"Composite"},
    {"state":"Jharkhand","district":"Hazaribagh","city":"Hazaribagh","climate":"Composite"},
    {"state":"Jharkhand","district":"Jamtara","city":"Jamtara","climate":"Composite"},
    {"state":"Jharkhand","district":"Koderma","city":"Koderma","climate":"Composite"},
    {"state":"Jharkhand","district":"Latehar","city":"Latehar","climate":"Composite"},
    {"state":"Jharkhand","district":"Lohardaga","city":"Lohardaga","climate":"Composite"},
    {"state":"Jharkhand","district":"Pakur","city":"Pakur","climate":"Composite"},
    {"state":"Jharkhand","district":"Palamu","city":"Medininagar","climate":"Composite"},
    {"state":"Jharkhand","district":"Ramgarh","city":"Ramgarh Nagar Parishad","climate":"Composite"},
    {"state":"Jharkhand","district":"Ranchi","city":"Ranchi","climate":"Composite"},
    {"state":"Jharkhand","district":"Sahebganj","city":"Sahibganj","climate":"Composite"},
    {"state":"Jharkhand","district":"Saraikela - Kharswan","city":"Seraikela","climate":"Composite"},
    {"state":"Jharkhand","district":"Simdega","city":"Simdega","climate":"Composite"},
    {"state":"Jharkhand","district":"West Singhbhum","city":"Chaibasa","climate":"Composite"},
    {"state":"Karnataka","district":"Bagalkote","city":"Bagalkot","climate":"Composite"},
    {"state":"Karnataka","district":"Ballary","city":"Bellary","climate":"Hot and Dry"},
    {"state":"Karnataka","district":"Belagavi","city":"Belgaum","climate":"Composite"},
    {"state":"Karnataka","district":"Bengaluru Rural","city":"Devanahalli","climate":"Composite"},
    {"state":"Karnataka","district":"Bengaluru Urban","city":"Bruhat Bengaluru Mahanagara Palike","climate":"Composite"},
    {"state":"Karnataka","district":"Bidar","city":"Bidar","climate":"Composite"},
    {"state":"Karnataka","district":"Chamarajanagara","city":"Chamarajanagar","climate":"Composite"},
    {"state":"Karnataka","district":"Chikkaballapura","city":"Chikkaballapura","climate":"Composite"},
    {"state":"Karnataka","district":"Chikkamagaluru","city":"Chikmagalur","climate":"Composite"},
    {"state":"Karnataka","district":"Chitradurga","city":"Chitradurga","climate":"Composite"},
    {"state":"Karnataka","district":"Dakshina kannada","city":"Mangalore","climate":"Composite"},
    {"state":"Karnataka","district":"Davangere","city":"Davanagere","climate":"Composite"},
    {"state":"Karnataka","district":"Dharwada","city":"Hubli-Dharwad","climate":"Composite"},
    {"state":"Karnataka","district":"Gadag","city":"Gadag-Betigeri","climate":"Composite"},
    {"state":"Karnataka","district":"Hassan","city":"Hassan","climate":"Composite"},
    {"state":"Karnataka","district":"Haveri","city":"Haveri","climate":"Composite"},
    {"state":"Karnataka","district":"Kalaburagi","city":"Gulbarga","climate":"Hot and Dry"},
    {"state":"Karnataka","district":"Kodagu","city":"Madikeri","climate":"Composite"},
    {"state":"Karnataka","district":"Kolar","city":"Kolar","climate":"Composite"},
    {"state":"Karnataka","district":"Koppal","city":"Koppal","climate":"Composite"},
    {"state":"Karnataka","district":"Mandya","city":"Mandya","climate":"Composite"},
    {"state":"Karnataka","district":"Mysuru","city":"Mysore","climate":"Composite"},
    {"state":"Karnataka","district":"Raichur","city":"Raichur","climate":"Hot and Dry"},
    {"state":"Karnataka","district":"Ramanagara","city":"Ramanagara","climate":"Composite"},
    {"state":"Karnataka","district":"Shivamogga","city":"Shimoga","climate":"Composite"},
    {"state":"Karnataka","district":"Tumakuru","city":"Tumkur","climate":"Composite"},
    {"state":"Karnataka","district":"Udupi","city":"Udupi","climate":"Composite"},
    {"state":"Karnataka","district":"Uttara Kannada","city":"Karwar","climate":"Composite"},
    {"state":"Karnataka","district":"Vijayapura","city":"Bijapur","climate":"Hot and Dry"},
    {"state":"Karnataka","district":"Yadgir","city":"Yadgir","climate":"Hot and Dry"},
    {"state":"Kerala","district":"Alappuzha","city":"Alappuzha","climate":"Warm & humid"},
    {"state":"Kerala","district":"Ernakulam","city":"Kochi","climate":"Warm & humid"},
    {"state":"Kerala","district":"Idukki","city":"Thodupuzha","climate":"Warm & humid"},
    {"state":"Kerala","district":"Kannur","city":"Kannur","climate":"Warm & humid"},
    {"state":"Kerala","district":"Kannur","city":"Thalassery","climate":"Warm & humid"},
    {"state":"Kerala","district":"Kasaragod","city":"Kasaragod","climate":"Warm & humid"},
    {"state":"Kerala","district":"Kollam","city":"Kollam","climate":"Warm & humid"},
    {"state":"Kerala","district":"Kottayam","city":"Kottayam","climate":"Warm & humid"},
    {"state":"Kerala","district":"Kozhikode","city":"Kozhikode","climate":"Warm & humid"},
    {"state":"Kerala","district":"Malappuram","city":"Malappuram","climate":"Warm & humid"},
    {"state":"Kerala","district":"Palakkad","city":"Palakkad","climate":"Warm & humid"},
    {"state":"Kerala","district":"Pathanamthitta","city":"Pathanamthitta","climate":"Warm & humid"},
    {"state":"Kerala","district":"Thiruvananthapuram","city":"Thiruvananthapuram","climate":"Warm & humid"},
    {"state":"Kerala","district":"Thrissur","city":"Thrissur","climate":"Warm & humid"},
    {"state":"Kerala","district":"Wayanad","city":"Kalpetta","climate":"Warm & humid"},
    {"state":"Ladakh","district":"Kargil","city":"Kargil","climate":"Cold"},
    {"state":"Ladakh","district":"Leh","city":"Leh","climate":"Cold"},
    {"state":"Madhya Pradesh","district":"Agar Malwa","city":"Agar","climate":"Composite"},
    {"state":"Madhya Pradesh","district":"Alirajpur","city":"Alirajpur","climate":"Composite"},
    {"state":"Madhya Pradesh","district":"Anuppur","city":"Anuppur","climate":"Composite"},
    {"state":"Madhya Pradesh","district":"Ashoknagar","city":"Ashoknagar","climate":"Composite"},
    {"state":"Madhya Pradesh","district":"Balaghat","city":"Balaghat","climate":"Composite"},
    {"state":"Madhya Pradesh","district":"Barwani","city":"Badwani","climate":"Composite"},
    {"state":"Madhya Pradesh","district":"Betul","city":"Betul","climate":"Composite"},
    {"state":"Madhya Pradesh","district":"Bhind","city":"Bhind","climate":"Composite"},
    {"state":"Madhya Pradesh","district":"Bhopal","city":"Bhopal","climate":"Composite"},
    {"state":"Madhya Pradesh","district":"Burhanpur","city":"Burhanpur","climate":"Composite"},
    {"state":"Madhya Pradesh","district":"Chhatarpur","city":"Chhatarpur","climate":"Composite"},
    {"state":"Madhya Pradesh","district":"Chhindwara","city":"Chhindwara","climate":"Composite"},
    {"state":"Madhya Pradesh","district":"Damoh","city":"Damoh","climate":"Composite"},
    {"state":"Madhya Pradesh","district":"Datia","city":"Datia","climate":"Composite"},
    {"state":"Madhya Pradesh","district":"Dewas","city":"Dewas","climate":"Composite"},
    {"state":"Madhya Pradesh","district":"Dhar","city":"Dhar","climate":"Composite"},
    {"state":"Madhya Pradesh","district":"Guna","city":"Guna","climate":"Composite"},
    {"state":"Madhya Pradesh","district":"Gwalior","city":"Gwalior","climate":"Composite"},
    {"state":"Madhya Pradesh","district":"Harda","city":"Harda","climate":"Composite"},
    {"state":"Madhya Pradesh","district":"Indore","city":"Indore","climate":"Composite"},
    {"state":"Madhya Pradesh","district":"Jabalpur","city":"Jabalpur","climate":"Composite"},
    {"state":"Madhya Pradesh","district":"Jhabua","city":"Jhabua","climate":"Composite"},
    {"state":"Madhya Pradesh","district":"Khandwa","city":"Khandwa","climate":"Composite"},
    {"state":"Madhya Pradesh","district":"Khargaon","city":"Khargone","climate":"Composite"},
    {"state":"Madhya Pradesh","district":"Mandla","city":"Mandla","climate":"Composite"},
    {"state":"Madhya Pradesh","district":"Mandsaur","city":"Mandsaur","climate":"Composite"},
    {"state":"Madhya Pradesh","district":"Morena","city":"Morena","climate":"Composite"},
    {"state":"Madhya Pradesh","district":"Murwara (Katni)","city":"Katni","climate":"Composite"},
    {"state":"Madhya Pradesh","district":"Narmadapuram","city":"Narmadapuram","climate":"Composite"},
    {"state":"Madhya Pradesh","district":"Narsinghpur","city":"Narsinghpur","climate":"Composite"},
    {"state":"Madhya Pradesh","district":"Neemuch","city":"Neemuch","climate":"Composite"},
    {"state":"Madhya Pradesh","district":"Panna","city":"Panna","climate":"Composite"},
    {"state":"Madhya Pradesh","district":"Raisen","city":"Raisen","climate":"Composite"},
    {"state":"Madhya Pradesh","district":"Rajgarh","city":"Rajgarh","climate":"Composite"},
    {"state":"Madhya Pradesh","district":"Ratlam","city":"Ratlam","climate":"Composite"},
    {"state":"Madhya Pradesh","district":"Rewa","city":"Rewa","climate":"Composite"},
    {"state":"Madhya Pradesh","district":"Sagar","city":"Sagar","climate":"Composite"},
    {"state":"Madhya Pradesh","district":"Satna","city":"Satna","climate":"Composite"},
    {"state":"Madhya Pradesh","district":"Sehore","city":"Sehore","climate":"Composite"},
    {"state":"Madhya Pradesh","district":"Seoni","city":"Seoni","climate":"Composite"},
    {"state":"Madhya Pradesh","district":"Shahdol","city":"Shahdol","climate":"Composite"},
    {"state":"Madhya Pradesh","district":"Shajapur","city":"Shajapur","climate":"Composite"},
    {"state":"Madhya Pradesh","district":"Sheopur","city":"Sheopur","climate":"Composite"},
    {"state":"Madhya Pradesh","district":"Shivpuri","city":"Shivpuri","climate":"Composite"},
    {"state":"Madhya Pradesh","district":"Sidhi","city":"Sidhi","climate":"Composite"},
    {"state":"Madhya Pradesh","district":"Singrauli","city":"Singrauli","climate":"Composite"},
    {"state":"Madhya Pradesh","district":"Tikamgarh","city":"Tikamgarh","climate":"Composite"},
    {"state":"Madhya Pradesh","district":"Ujjain","city":"Ujjain","climate":"Composite"},
    {"state":"Madhya Pradesh","district":"Umaria","city":"Umaria","climate":"Composite"},
    {"state":"Madhya Pradesh","district":"Vidisha","city":"Vidisha","climate":"Composite"},
    {"state":"Maharashtra","district":"Ahmednagar","city":"Ahmednagar","climate":"Composite"},
    {"state":"Maharashtra","district":"Akola","city":"Akola","climate":"Composite"},
    {"state":"Maharashtra","district":"Amravati","city":"Amravati","climate":"Composite"},
    {"state":"Maharashtra","district":"Aurangabad","city":"Aurangabad","climate":"Hot and Dry"},
    {"state":"Maharashtra","district":"Beed","city":"Beed","climate":"Composite"},
    {"state":"Maharashtra","district":"Bhandara","city":"Bhandara","climate":"Composite"},
    {"state":"Maharashtra","district":"Buldana","city":"Buldana","climate":"Composite"},
    {"state":"Maharashtra","district":"Chandrapur","city":"Chandrapur","climate":"Composite"},
    {"state":"Maharashtra","district":"Dhule","city":"Dhule","climate":"Composite"},
    {"state":"Maharashtra","district":"Gadchiroli","city":"Gadchiroli","climate":"Composite"},
    {"state":"Maharashtra","district":"Gondiya","city":"Gondiya","climate":"Composite"},
    {"state":"Maharashtra","district":"Hingoli","city":"Hingoli","climate":"Composite"},
    {"state":"Maharashtra","district":"Jalgaon","city":"Jalgaon","climate":"Composite"},
    {"state":"Maharashtra","district":"Jalna","city":"Jalna","climate":"Composite"},
    {"state":"Maharashtra","district":"Kolhapur","city":"Kolhapur","climate":"Composite"},
    {"state":"Maharashtra","district":"Latur","city":"Latur","climate":"Hot and Dry"},
    {"state":"Maharashtra","district":"Mumbai","city":"Greater Mumbai","climate":"Composite"},
    {"state":"Maharashtra","district":"Nagpur","city":"Nagpur","climate":"Composite"},
    {"state":"Maharashtra","district":"Nanded Waghala","city":"Nanded Waghala","climate":"Composite"},
    {"state":"Maharashtra","district":"Nandurbar","city":"Nandurbar","climate":"Composite"},
    {"state":"Maharashtra","district":"Nashik","city":"Nashik","climate":"Composite"},
    {"state":"Maharashtra","district":"Osmanabad","city":"Osmanabad","climate":"Hot and Dry"},
    {"state":"Maharashtra","district":"Palghar","city":"Palghar","climate":"Composite"},
    {"state":"Maharashtra","district":"Parbhani","city":"Parbhani","climate":"Composite"},
    {"state":"Maharashtra","district":"Pune","city":"Pune","climate":"Composite"},
    {"state":"Maharashtra","district":"Raigadh","city":"Alibag","climate":"Composite"},
    {"state":"Maharashtra","district":"Ratnagiri","city":"Ratnagiri","climate":"Composite"},
    {"state":"Maharashtra","district":"Sangli","city":"Sangli","climate":"Composite"},
    {"state":"Maharashtra","district":"Satara","city":"Satara","climate":"Composite"},
    {"state":"Maharashtra","district":"Sindhudurga","city":"Kudal","climate":"Composite"},
    {"state":"Maharashtra","district":"Solapur","city":"Solapur","climate":"Hot and Dry"},
    {"state":"Maharashtra","district":"Thane","city":"Thane","climate":"Composite"},
    {"state":"Maharashtra","district":"Wardha","city":"Wardha","climate":"Composite"},
    {"state":"Maharashtra","district":"Washim","city":"Washim","climate":"Composite"},
    {"state":"Maharashtra","district":"Yavatmal","city":"Yavatmal","climate":"Composite"},
    {"state":"Manipur","district":"Bishnupur","city":"Bishnupur","climate":"Warm & humid"},
    {"state":"Manipur","district":"Imphal East","city":"Imphal","climate":"Warm & humid"},
    {"state":"Manipur","district":"Thoubal","city":"Thoubal","climate":"Warm & humid"},
    {"state":"Meghalaya","district":"East Garo Hills","city":"Williamnagar","climate":"Warm & humid"},
    {"state":"Meghalaya","district":"East Khasi","city":"Shillong","climate":"Warm & humid"},
    {"state":"Meghalaya","district":"West Garo Hills","city":"Tura","climate":"Warm & humid"},
    {"state":"Mizoram","district":"Aizawl","city":"Aizawl","climate":"Warm & humid"},
    {"state":"Mizoram","district":"Champhai","city":"Champhai","climate":"Warm & humid"},
    {"state":"Mizoram","district":"Kolasib","city":"Kolasib","climate":"Warm & humid"},
    {"state":"Mizoram","district":"Lunglei","city":"Lunglei","climate":"Warm & humid"},
    {"state":"Nagaland","district":"Dimapur","city":"Dimapur","climate":"Warm & humid"},
    {"state":"Nagaland","district":"Kohima","city":"Kohima","climate":"Warm & humid"},
    {"state":"Nagaland","district":"Mokokchung","city":"Mokokchung","climate":"Warm & humid"},
    {"state":"Nagaland","district":"Mon","city":"Mon","climate":"Warm & humid"},
    {"state":"Nagaland","district":"Phek","city":"Phek","climate":"Warm & humid"},
    {"state":"Nagaland","district":"Tuensang","city":"Tuensang","climate":"Warm & humid"},
    {"state":"Nagaland","district":"Wokha","city":"Wokha","climate":"Warm & humid"},
    {"state":"Odisha","district":"Anugul","city":"Anugul","climate":"Warm & humid"},
    {"state":"Odisha","district":"Balangir","city":"Balangir","climate":"Warm & humid"},
    {"state":"Odisha","district":"Baleshwar","city":"Baleshwar Town","climate":"Warm & humid"},
    {"state":"Odisha","district":"Bargarh","city":"Bargarh","climate":"Warm & humid"},
    {"state":"Odisha","district":"Bhadrak","city":"Bhadrak","climate":"Warm & humid"},
    {"state":"Odisha","district":"Cuttack","city":"Cuttack","climate":"Warm & humid"},
    {"state":"Odisha","district":"Dhenkanal","city":"Dhenkanal","climate":"Warm & humid"},
    {"state":"Odisha","district":"Gajapati","city":"Paralakhemundi","climate":"Warm & humid"},
    {"state":"Odisha","district":"Ganjam","city":"Brahmapur","climate":"Warm & humid"},
    {"state":"Odisha","district":"Jagatsinghpur","city":"Jagatsinghapur","climate":"Warm & humid"},
    {"state":"Odisha","district":"Jajpur","city":"Jajapur","climate":"Warm & humid"},
    {"state":"Odisha","district":"Jharsuguda","city":"Jharsuguda","climate":"Warm & humid"},
    {"state":"Odisha","district":"Kalahandi","city":"Bhawanipatna","climate":"Warm & humid"},
    {"state":"Odisha","district":"Kendrapara","city":"Kendrapara","climate":"Warm & humid"},
    {"state":"Odisha","district":"Keonjhar","city":"Kendujhar","climate":"Warm & humid"},
    {"state":"Odisha","district":"Khurda","city":"Bhubaneswar","climate":"Warm & humid"},
    {"state":"Odisha","district":"Koraput","city":"Koraput","climate":"Warm & humid"},
    {"state":"Odisha","district":"Malkangiri","city":"Malkangiri","climate":"Warm & humid"},
    {"state":"Odisha","district":"Mayurbhanj","city":"Baripada Town","climate":"Warm & humid"},
    {"state":"Odisha","district":"Nabarangapur","city":"Nabarangapur","climate":"Warm & humid"},
    {"state":"Odisha","district":"Nayagarh","city":"Nayagarh","climate":"Warm & humid"},
    {"state":"Odisha","district":"Puri","city":"Puri","climate":"Warm & humid"},
    {"state":"Odisha","district":"Rayagada","city":"Rayagada","climate":"Warm & humid"},
    {"state":"Odisha","district":"Sambalpur","city":"Sambalpur Town","climate":"Warm & humid"},
    {"state":"Odisha","district":"Sundergarh","city":"Raurkela Town","climate":"Warm & humid"},
    {"state":"Puducherry","district":"Karaikal","city":"Karaikal","climate":"Warm & humid"},
    {"state":"Puducherry","district":"Mahe","city":"Mahe","climate":"Warm & humid"},
    {"state":"Puducherry","district":"Puducherry","city":"Puducherry","climate":"Warm & humid"},
    {"state":"Punjab","district":"Amritsar","city":"Amritsar","climate":"Composite"},
    {"state":"Punjab","district":"Barnala","city":"Barnala","climate":"Composite"},
    {"state":"Punjab","district":"Bathinda","city":"Bathinda","climate":"Composite"},
    {"state":"Punjab","district":"Faridkot","city":"Faridkot","climate":"Composite"},
    {"state":"Punjab","district":"Fazilka","city":"Abohar","climate":"Composite"},
    {"state":"Punjab","district":"Firozpur","city":"Firozpur","climate":"Composite"},
    {"state":"Punjab","district":"Gurdaspur","city":"Gurdaspur","climate":"Composite"},
    {"state":"Punjab","district":"Hoshiarpur","city":"Hoshiarpur","climate":"Composite"},
    {"state":"Punjab","district":"Jalandhar","city":"Jalandhar","climate":"Composite"},
    {"state":"Punjab","district":"Kapurthala","city":"Kapurthala","climate":"Composite"},
    {"state":"Punjab","district":"Ludhiana","city":"Ludhiana","climate":"Composite"},
    {"state":"Punjab","district":"Mansa","city":"Mansa","climate":"Composite"},
    {"state":"Punjab","district":"Moga","city":"Moga","climate":"Composite"},
    {"state":"Punjab","district":"Muktsar","city":"Muktsar","climate":"Composite"},
    {"state":"Punjab","district":"Pathankot","city":"Pathankot","climate":"Composite"},
    {"state":"Punjab","district":"Patiala","city":"Patiala","climate":"Composite"},
    {"state":"Punjab","district":"Rupnagar","city":"Rupnagar","climate":"Composite"},
    {"state":"Punjab","district":"S.A.S. Nagar (Mohali)","city":"Mohali","climate":"Composite"},
    {"state":"Punjab","district":"Sangrur","city":"Sangrur","climate":"Composite"},
    {"state":"Punjab","district":"Sirhind Fatehgarh Sahib","city":"Sirhind Fatehgarh Sahib","climate":"Composite"},
    {"state":"Punjab","district":"Tarn Taran","city":"Tarn Taran","climate":"Composite"},
    {"state":"Rajasthan","district":"Ajmer","city":"Ajmer","climate":"Hot and Dry"},
    {"state":"Rajasthan","district":"Alwar","city":"Alwar","climate":"Hot and Dry"},
    {"state":"Rajasthan","district":"Banswara","city":"Banswara","climate":"Hot and Dry"},
    {"state":"Rajasthan","district":"Baran","city":"Baran","climate":"Hot and Dry"},
    {"state":"Rajasthan","district":"Barmer","city":"Barmer","climate":"Hot and Dry"},
    {"state":"Rajasthan","district":"Bharatpur","city":"Bharatpur","climate":"Hot and Dry"},
    {"state":"Rajasthan","district":"Bhilwara","city":"Bhilwara","climate":"Hot and Dry"},
    {"state":"Rajasthan","district":"Bikaner","city":"Bikaner","climate":"Hot and Dry"},
    {"state":"Rajasthan","district":"Bundi","city":"Bundi","climate":"Hot and Dry"},
    {"state":"Rajasthan","district":"Chittaurgarh","city":"Chittaurgarh","climate":"Hot and Dry"},
    {"state":"Rajasthan","district":"Churu","city":"Churu","climate":"Hot and Dry"},
    {"state":"Rajasthan","district":"Dausa","city":"Dausa","climate":"Hot and Dry"},
    {"state":"Rajasthan","district":"Dhaulpur","city":"Dhaulpur","climate":"Hot and Dry"},
    {"state":"Rajasthan","district":"Dungarpur","city":"Dungarpur","climate":"Hot and Dry"},
    {"state":"Rajasthan","district":"Ganganagar","city":"Ganganagar","climate":"Hot and Dry"},
    {"state":"Rajasthan","district":"Hanumangarh","city":"Hanumangarh","climate":"Hot and Dry"},
    {"state":"Rajasthan","district":"Jaipur","city":"Jaipur Greater","climate":"Hot and Dry"},
    {"state":"Rajasthan","district":"Jaisalmer","city":"Jaisalmer","climate":"Hot and Dry"},
    {"state":"Rajasthan","district":"Jalor","city":"Jalor","climate":"Hot and Dry"},
    {"state":"Rajasthan","district":"Jhalawar","city":"Jhalawar","climate":"Hot and Dry"},
    {"state":"Rajasthan","district":"Jhunjhunun","city":"Jhunjhunun","climate":"Hot and Dry"},
    {"state":"Rajasthan","district":"Jodhpur","city":"Jodhpur North","climate":"Hot and Dry"},
    {"state":"Rajasthan","district":"Karauli","city":"Karauli","climate":"Hot and Dry"},
    {"state":"Rajasthan","district":"Kota","city":"Kota North","climate":"Hot and Dry"},
    {"state":"Rajasthan","district":"Nagaur","city":"Nagaur","climate":"Hot and Dry"},
    {"state":"Rajasthan","district":"Pali","city":"Pali","climate":"Hot and Dry"},
    {"state":"Rajasthan","district":"Pratapgarh","city":"Pratapgarh","climate":"Hot and Dry"},
    {"state":"Rajasthan","district":"Rajsamand","city":"Rajsamand","climate":"Hot and Dry"},
    {"state":"Rajasthan","district":"Sawai Madhopur","city":"Sawai Madhopur","climate":"Hot and Dry"},
    {"state":"Rajasthan","district":"Sikar","city":"Sikar","climate":"Hot and Dry"},
    {"state":"Rajasthan","district":"Sirohi","city":"Sirohi","climate":"Hot and Dry"},
    {"state":"Rajasthan","district":"Tonk","city":"Tonk","climate":"Hot and Dry"},
    {"state":"Rajasthan","district":"Udaipur","city":"Udaipur","climate":"Hot and Dry"},
    {"state":"Sikkim","district":"Gangtok","city":"Gangtok","climate":"Cold"},
    {"state":"Sikkim","district":"Namchi","city":"Namchi","climate":"Cold"},
    {"state":"Tamil Nadu","district":"Ariyalur","city":"Ariyalur","climate":"Warm & humid"},
    {"state":"Tamil Nadu","district":"Chengalpattu","city":"Chengalpattu","climate":"Warm & humid"},
    {"state":"Tamil Nadu","district":"Chennai","city":"Chennai","climate":"Warm & humid"},
    {"state":"Tamil Nadu","district":"Coimbatore","city":"Coimbatore","climate":"Warm & humid"},
    {"state":"Tamil Nadu","district":"Cuddalore","city":"Cuddalore","climate":"Warm & humid"},
    {"state":"Tamil Nadu","district":"Dharmapuri","city":"Dharmapuri","climate":"Warm & humid"},
    {"state":"Tamil Nadu","district":"Dindigul","city":"Dindigul","climate":"Warm & humid"},
    {"state":"Tamil Nadu","district":"Erode","city":"Erode","climate":"Warm & humid"},
    {"state":"Tamil Nadu","district":"Kallakurichi","city":"Kallakurichi","climate":"Warm & humid"},
    {"state":"Tamil Nadu","district":"Kancheepuram","city":"Kancheepuram","climate":"Warm & humid"},
    {"state":"Tamil Nadu","district":"Kanniyakumari","city":"Nagarcoil","climate":"Warm & humid"},
    {"state":"Tamil Nadu","district":"Karur","city":"Karur","climate":"Warm & humid"},
    {"state":"Tamil Nadu","district":"Krishnagiri","city":"Krishnagiri","climate":"Warm & humid"},
    {"state":"Tamil Nadu","district":"Madurai","city":"Madurai","climate":"Warm & humid"},
    {"state":"Tamil Nadu","district":"Mayiladuthurai","city":"Mayiladuthurai","climate":"Warm & humid"},
    {"state":"Tamil Nadu","district":"Nagapattinam","city":"Nagapattinam","climate":"Warm & humid"},
    {"state":"Tamil Nadu","district":"Namakkal","city":"Namakkal","climate":"Warm & humid"},
    {"state":"Tamil Nadu","district":"Perambalur","city":"Perambalur","climate":"Warm & humid"},
    {"state":"Tamil Nadu","district":"Pudukottai","city":"Pudukottai","climate":"Warm & humid"},
    {"state":"Tamil Nadu","district":"Ramanathapuram","city":"Ramanathapuram","climate":"Warm & humid"},
    {"state":"Tamil Nadu","district":"Ranipet","city":"Ranipet","climate":"Warm & humid"},
    {"state":"Tamil Nadu","district":"Salem","city":"Salem","climate":"Warm & humid"},
    {"state":"Tamil Nadu","district":"Sivaganga","city":"Sivagangai","climate":"Warm & humid"},
    {"state":"Tamil Nadu","district":"Tenkasi","city":"Tenkasi","climate":"Warm & humid"},
    {"state":"Tamil Nadu","district":"Thanjavur","city":"Thanjavur","climate":"Warm & humid"},
    {"state":"Tamil Nadu","district":"The Nilgiris","city":"Udagamandalam","climate":"Warm & humid"},
    {"state":"Tamil Nadu","district":"Theni","city":"Theni Alinagaram","climate":"Warm & humid"},
    {"state":"Tamil Nadu","district":"Thirupathur","city":"Tirupathur","climate":"Warm & humid"},
    {"state":"Tamil Nadu","district":"Thiruvallur","city":"Tiruvallur","climate":"Warm & humid"},
    {"state":"Tamil Nadu","district":"Thoothukkudi","city":"Thoothukudi Corporation","climate":"Warm & humid"},
    {"state":"Tamil Nadu","district":"Tirunelveli","city":"Tirunelveli","climate":"Warm & humid"},
    {"state":"Tamil Nadu","district":"Tiruppurur","city":"Tirupppur","climate":"Warm & humid"},
    {"state":"Tamil Nadu","district":"Tiruvannamalai","city":"Tiruvannamalai","climate":"Warm & humid"},
    {"state":"Tamil Nadu","district":"Tiruvarur","city":"Thiruvarur","climate":"Warm & humid"},
    {"state":"Tamil Nadu","district":"Trichy","city":"Tiruchirapalli","climate":"Warm & humid"},
    {"state":"Tamil Nadu","district":"Vellore","city":"Vellore","climate":"Warm & humid"},
    {"state":"Tamil Nadu","district":"Viluppuram","city":"Villupuram","climate":"Warm & humid"},
    {"state":"Tamil Nadu","district":"Virudhunagar","city":"Virudhunagar","climate":"Warm & humid"},
    {"state":"Telangana","district":"Adilabad","city":"Adilabad","climate":"Hot and Dry"},
    {"state":"Telangana","district":"Bhadradri Kothagudem","city":"Kothagudem","climate":"Hot and Dry"},
    {"state":"Telangana","district":"Hanumakonda","city":"Warangal","climate":"Hot and Dry"},
    {"state":"Telangana","district":"Hyderabad","city":"Greater Hyderabad","climate":"Hot and Dry"},
    {"state":"Telangana","district":"Jagitial","city":"Jagitial","climate":"Hot and Dry"},
    {"state":"Telangana","district":"Jangaon","city":"Jangaon","climate":"Hot and Dry"},
    {"state":"Telangana","district":"Karimnagar","city":"Karimnagar","climate":"Hot and Dry"},
    {"state":"Telangana","district":"Khammam","city":"Khammam","climate":"Hot and Dry"},
    {"state":"Telangana","district":"Mahabubabad","city":"Mahaboobabad","climate":"Hot and Dry"},
    {"state":"Telangana","district":"Mahabubnagar","city":"Mahabubnagar","climate":"Hot and Dry"},
    {"state":"Telangana","district":"Mancherial","city":"Mancherial","climate":"Hot and Dry"},
    {"state":"Telangana","district":"Medak","city":"Medak","climate":"Hot and Dry"},
    {"state":"Telangana","district":"Nalgonda","city":"Nalgonda","climate":"Hot and Dry"},
    {"state":"Telangana","district":"Nirmal","city":"Nirmal","climate":"Hot and Dry"},
    {"state":"Telangana","district":"Nizamabad","city":"Nizamabad","climate":"Hot and Dry"},
    {"state":"Telangana","district":"Peddapally","city":"Ramagundam","climate":"Hot and Dry"},
    {"state":"Telangana","district":"Rangareddy","city":"Shadnagar","climate":"Hot and Dry"},
    {"state":"Telangana","district":"Sangareddy","city":"Sangareddy","climate":"Hot and Dry"},
    {"state":"Telangana","district":"Siddipet","city":"Siddipet","climate":"Hot and Dry"},
    {"state":"Telangana","district":"Suryapet","city":"Suryapet","climate":"Hot and Dry"},
    {"state":"Telangana","district":"Vikarabad","city":"Vikarabad","climate":"Hot and Dry"},
    {"state":"Telangana","district":"Wanapathy","city":"Wanaparthy","climate":"Hot and Dry"},
    {"state":"Telangana","district":"Warangal","city":"Narsampet","climate":"Hot and Dry"},
    {"state":"Telangana","district":"Yadadri Bhuvanagiri","city":"Bhongir","climate":"Hot and Dry"},
    {"state":"Tripura","district":"Dhalai","city":"Ambassa","climate":"Warm & humid"},
    {"state":"Tripura","district":"Gomati","city":"Udaipur","climate":"Warm & humid"},
    {"state":"Tripura","district":"North Tripura","city":"Dharmanagar","climate":"Warm & humid"},
    {"state":"Tripura","district":"South Tripura","city":"Belonia","climate":"Warm & humid"},
    {"state":"Tripura","district":"West Tripura","city":"Agartala","climate":"Warm & humid"},
    {"state":"Uttar Pradesh","district":"Agra","city":"Agra","climate":"Composite"},
    {"state":"Uttar Pradesh","district":"Aligarh","city":"Aligarh","climate":"Composite"},
    {"state":"Uttar Pradesh","district":"Ambedkar Nagar","city":"Akbarpur","climate":"Composite"},
    {"state":"Uttar Pradesh","district":"Amroha","city":"Amroha","climate":"Composite"},
    {"state":"Uttar Pradesh","district":"Auraiya","city":"Auraiya","climate":"Composite"},
    {"state":"Uttar Pradesh","district":"Ayodhya","city":"Ayodhya","climate":"Composite"},
    {"state":"Uttar Pradesh","district":"Azamgarh","city":"Azamgarh","climate":"Composite"},
    {"state":"Uttar Pradesh","district":"Baghpat","city":"Baghpat","climate":"Composite"},
    {"state":"Uttar Pradesh","district":"Bahraich","city":"Bahraich","climate":"Composite"},
    {"state":"Uttar Pradesh","district":"Ballia","city":"Ballia","climate":"Composite"},
    {"state":"Uttar Pradesh","district":"Balrampur","city":"Balrampur","climate":"Composite"},
    {"state":"Uttar Pradesh","district":"Banda","city":"Banda","climate":"Composite"},
    {"state":"Uttar Pradesh","district":"Barabanki","city":"Barabanki","climate":"Composite"},
    {"state":"Uttar Pradesh","district":"Bareilly","city":"Bareilly","climate":"Composite"},
    {"state":"Uttar Pradesh","district":"Basti","city":"Basti","climate":"Composite"},
    {"state":"Uttar Pradesh","district":"Bhadohi","city":"Bhadohi","climate":"Composite"},
    {"state":"Uttar Pradesh","district":"Bijnor","city":"Bijnor","climate":"Composite"},
    {"state":"Uttar Pradesh","district":"Budaun","city":"Budaun","climate":"Composite"},
    {"state":"Uttar Pradesh","district":"Bulandshahar","city":"Bulandshahr","climate":"Composite"},
    {"state":"Uttar Pradesh","district":"Chandauli","city":"Chandauli","climate":"Composite"},
    {"state":"Uttar Pradesh","district":"Chitrakoot","city":"Chitrakoot Dham","climate":"Composite"},
    {"state":"Uttar Pradesh","district":"Deoria","city":"Deoria","climate":"Composite"},
    {"state":"Uttar Pradesh","district":"Etah","city":"Etah","climate":"Composite"},
    {"state":"Uttar Pradesh","district":"Etawah","city":"Etawah","climate":"Composite"},
    {"state":"Uttar Pradesh","district":"Farrukhabad","city":"Farrukhabad-Cum-Fatehgarh","climate":"Composite"},
    {"state":"Uttar Pradesh","district":"Fatehpur","city":"Fatehpur","climate":"Composite"},
    {"state":"Uttar Pradesh","district":"Firozabad","city":"Firozabad","climate":"Composite"},
    {"state":"Uttar Pradesh","district":"Gautam Buddha Nagar","city":"Noida","climate":"Composite"},
    {"state":"Uttar Pradesh","district":"Ghaziabad","city":"Ghaziabad","climate":"Composite"},
    {"state":"Uttar Pradesh","district":"Ghazipur","city":"Ghazipur","climate":"Composite"},
    {"state":"Uttar Pradesh","district":"Gonda","city":"Gonda","climate":"Composite"},
    {"state":"Uttar Pradesh","district":"Gorakhpur","city":"Gorakhpur","climate":"Composite"},
    {"state":"Uttar Pradesh","district":"Hamirpur","city":"Hamirpur","climate":"Composite"},
    {"state":"Uttar Pradesh","district":"Hapur","city":"Hapur","climate":"Composite"},
    {"state":"Uttar Pradesh","district":"Hardoi","city":"Hardoi","climate":"Composite"},
    {"state":"Uttar Pradesh","district":"Hathras","city":"Hathras","climate":"Composite"},
    {"state":"Uttar Pradesh","district":"Jalaun","city":"Jalaun","climate":"Composite"},
    {"state":"Uttar Pradesh","district":"Jaunpur","city":"Jaunpur","climate":"Composite"},
    {"state":"Uttar Pradesh","district":"Jhansi","city":"Jhansi","climate":"Composite"},
    {"state":"Uttar Pradesh","district":"Kannauj","city":"Kannauj","climate":"Composite"},
    {"state":"Uttar Pradesh","district":"Kanpur","city":"Kanpur","climate":"Composite"},
    {"state":"Uttar Pradesh","district":"Kanpur Dehat","city":"Akbarpur","climate":"Composite"},
    {"state":"Uttar Pradesh","district":"Kasganj","city":"Kasganj","climate":"Composite"},
    {"state":"Uttar Pradesh","district":"Kaushambi","city":"Kaushambi","climate":"Composite"},
    {"state":"Uttar Pradesh","district":"Kheeri","city":"Lakhimpur","climate":"Composite"},
    {"state":"Uttar Pradesh","district":"Kushinagar","city":"Kushinagar","climate":"Composite"},
    {"state":"Uttar Pradesh","district":"Lalitpur","city":"Lalitpur","climate":"Composite"},
    {"state":"Uttar Pradesh","district":"Lucknow","city":"Lucknow","climate":"Composite"},
    {"state":"Uttar Pradesh","district":"Maharajganj","city":"Maharajganj","climate":"Composite"},
    {"state":"Uttar Pradesh","district":"Mahoba","city":"Mahoba","climate":"Composite"},
    {"state":"Uttar Pradesh","district":"Mainpuri","city":"Mainpuri","climate":"Composite"},
    {"state":"Uttar Pradesh","district":"Mathura","city":"Mathura-Vrindavan","climate":"Composite"},
    {"state":"Uttar Pradesh","district":"Mau","city":"Maunath Bhanjan","climate":"Composite"},
    {"state":"Uttar Pradesh","district":"Meerut","city":"Meerut","climate":"Composite"},
    {"state":"Uttar Pradesh","district":"Mirzapur-Cum-Vindhyachal","city":"Mirzapur-Cum-Vindhyachal","climate":"Composite"},
    {"state":"Uttar Pradesh","district":"Moradabad","city":"Moradabad","climate":"Composite"},
    {"state":"Uttar Pradesh","district":"Muzaffarnagar","city":"Muzaffarnagar","climate":"Composite"},
    {"state":"Uttar Pradesh","district":"Pilibhit","city":"Pilibhit","climate":"Composite"},
    {"state":"Uttar Pradesh","district":"Pratapgarh","city":"Pratapgarh City","climate":"Composite"},
    {"state":"Uttar Pradesh","district":"Prayagraj","city":"Prayagraj","climate":"Composite"},
    {"state":"Uttar Pradesh","district":"Raebareli","city":"Rae Bareli","climate":"Composite"},
    {"state":"Uttar Pradesh","district":"Rampur","city":"Rampur","climate":"Composite"},
    {"state":"Uttar Pradesh","district":"Saharanpur","city":"Saharanpur","climate":"Composite"},
    {"state":"Uttar Pradesh","district":"Sambhal","city":"Sambhal","climate":"Composite"},
    {"state":"Uttar Pradesh","district":"Sant Kabir Nagar","city":"Khalilabad","climate":"Composite"},
    {"state":"Uttar Pradesh","district":"Shahjahanpur","city":"Shahjahanpur","climate":"Composite"},
    {"state":"Uttar Pradesh","district":"Shamli","city":"Shamli","climate":"Composite"},
    {"state":"Uttar Pradesh","district":"Siddharthnagar","city":"Siddharthnagar","climate":"Composite"},
    {"state":"Uttar Pradesh","district":"Sitapur","city":"Sitapur","climate":"Composite"},
    {"state":"Uttar Pradesh","district":"Sonbhadra","city":"Sonbhadra","climate":"Composite"},
    {"state":"Uttar Pradesh","district":"Sultanpur","city":"Sultanpur","climate":"Composite"},
    {"state":"Uttar Pradesh","district":"Unnao","city":"Unnao","climate":"Composite"},
    {"state":"Uttar Pradesh","district":"Varanasi","city":"Varanasi","climate":"Composite"},
    {"state":"Uttarakhand","district":"Almora","city":"Almora","climate":"Temperate"},
    {"state":"Uttarakhand","district":"Bageshwar","city":"Bageshwar","climate":"Temperate"},
    {"state":"Uttarakhand","district":"Chamoli Gopeshwar","city":"Chamoli-Gopeshwar","climate":"Temperate"},
    {"state":"Uttarakhand","district":"Champawat","city":"Champawat","climate":"Temperate"},
    {"state":"Uttarakhand","district":"Dehradun","city":"Dehradun","climate":"Composite"},
    {"state":"Uttarakhand","district":"Hardwar","city":"Hardwar","climate":"Composite"},
    {"state":"Uttarakhand","district":"Nainital","city":"Haldwani","climate":"Temperate"},
    {"state":"Uttarakhand","district":"Nainital","city":"Nainital","climate":"Temperate"},
    {"state":"Uttarakhand","district":"Pauri Garhwal","city":"Kotdwara","climate":"Temperate"},
    {"state":"Uttarakhand","district":"Pithoragarh","city":"Pithoragarh","climate":"Temperate"},
    {"state":"Uttarakhand","district":"Rudraprayag","city":"Rudraprayag","climate":"Temperate"},
    {"state":"Uttarakhand","district":"Tehri","city":"Tehri","climate":"Temperate"},
    {"state":"Uttarakhand","district":"Udhamsingh Nagar","city":"Rudrapur","climate":"Composite"},
    {"state":"Uttarakhand","district":"Uttarkashi","city":"Uttarkashi","climate":"Temperate"},
    {"state":"West Bengal","district":"Bankura","city":"Bankura","climate":"Warm & humid"},
    {"state":"West Bengal","district":"Barddhamam","city":"Asansol","climate":"Warm & humid"},
    {"state":"West Bengal","district":"Barddhamam","city":"Barddhaman","climate":"Warm & humid"},
    {"state":"West Bengal","district":"Birbhum","city":"Suri","climate":"Warm & humid"},
    {"state":"West Bengal","district":"Dakshin Dinajpur","city":"Balurghat","climate":"Warm & humid"},
    {"state":"West Bengal","district":"Darjeeling","city":"Darjiling","climate":"Cold"},
    {"state":"West Bengal","district":"Darjeeling","city":"Siliguri","climate":"Cold"},
    {"state":"West Bengal","district":"Hooghly","city":"Chandannagar","climate":"Warm & humid"},
    {"state":"West Bengal","district":"Hooghly","city":"Serampore","climate":"Warm & humid"},
    {"state":"West Bengal","district":"Howrah","city":"Haora","climate":"Warm & humid"},
    {"state":"West Bengal","district":"Jalpaiguri","city":"Jalpaiguri","climate":"Warm & humid"},
    {"state":"West Bengal","district":"Koch Bihar","city":"Koch Bihar","climate":"Warm & humid"},
    {"state":"West Bengal","district":"Kolkata","city":"Kolkata","climate":"Warm & humid"},
    {"state":"West Bengal","district":"Maldah","city":"English Bazar","climate":"Warm & humid"},
    {"state":"West Bengal","district":"Murshidabad","city":"Berhampore","climate":"Warm & humid"},
    {"state":"West Bengal","district":"Nadia","city":"Krishnanagar","climate":"Warm & humid"},
    {"state":"West Bengal","district":"North 24 Parganas","city":"Barasat","climate":"Warm & humid"},
    {"state":"West Bengal","district":"North 24 Parganas","city":"Barrackpore","climate":"Warm & humid"},
    {"state":"West Bengal","district":"Paschim Medinipur","city":"Medinipur","climate":"Warm & humid"},
    {"state":"West Bengal","district":"Purba Midnapur","city":"Haldia","climate":"Warm & humid"},
    {"state":"West Bengal","district":"Puruliya","city":"Puruliya","climate":"Warm & humid"},
    {"state":"West Bengal","district":"South 24 Parganas","city":"Diamond Harbour","climate":"Warm & humid"},
    {"state":"West Bengal","district":"Uttar Dinajpur","city":"Raiganj","climate":"Warm & humid"},
    {"state":"Lakshadweep","district":"Lakshadweep","city":"Lakshadweep","climate":"Warm & humid"},
]


# ═══════════════════════════════════════════════════════════════════════════════
# CALCULATION ENGINE — mirrors Excel sheet-by-sheet logic
# ═══════════════════════════════════════════════════════════════════════════════

def ef_total(key):
    """Return total CO2e EF for a named fuel key."""
    e = EF[key]
    return e["co2"] + e["ch4"] + e["n2o"]


def calc_fuel_emission(activity_val, unit, ef_key):
    """
    Base Year GHG Inventory pattern:
      tCO2e = activity_tj * (CO2 + CH4_tCO2e + N2O_tCO2e)
    For electricity: activity_MWh * EF_t/MWh
    """
    e = EF[ef_key]
    if e["unit"] == "t/MWh":
        return activity_val * e["co2"]
    else:
        # already in TJ
        return activity_val * (e["co2"] + e["ch4"] + e["n2o"])


def fuel_to_tj(fuel_type, amount, input_unit):
    """Convert fuel quantity to TJ using ListsAndTables conversion factors."""
    c = FUEL_CONV.get(fuel_type, {})
    if input_unit == "kl" and "kl_to_tj" in c:
        return amount * c["kl_to_tj"]
    if input_unit == "tonne" and "t_to_tj" in c:
        return amount * c["t_to_tj"]
    if input_unit == "tj":
        return amount
    return 0.0


# ─── BUILDINGS & ENERGY ───────────────────────────────────────────────────────
def calc_buildings(d):
    """
    Mirrors: Base Year GHG Inventory → Building and Energy Sector
    Input units from form: TJ or MWh depending on fuel
    """
    subs = {}

    def sector_emit(prefix, fuels):
        total = 0
        for fuel_key, ef_key in fuels:
            val = float(d.get(f"{prefix}_{fuel_key}", 0) or 0)
            total += calc_fuel_emission(val, None, ef_key)
        return total

    subs["Residential"] = sector_emit("res", [
        ("Electricity", "Res_Electricity"),
        ("LPG",         "Res_LPG"),
        ("Firewood",    "Res_Firewood"),
        ("Kerosene",    "Res_Kerosene"),
        ("PNG",         "Res_PNG"),
        ("Coal",        "Res_Coal"),
    ])
    subs["Commercial"] = sector_emit("com", [
        ("Electricity", "Com_Electricity"),
        ("LPG",         "Com_LPG"),
        ("PNG",         "Com_PNG"),
        ("Firewood",    "Com_Firewood"),
        ("Kerosene",    "Com_Kerosene"),
    ])
    subs["Public & Institutional"] = sector_emit("ins", [
        ("Electricity", "Ins_Electricity"),
        ("LPG",         "Ins_LPG"),
    ])
    subs["Industrial"] = sector_emit("ind", [
        ("Electricity", "Ind_Electricity"),
        ("LPG",         "Ind_LPG"),
        ("Coal",        "Ind_Coal"),
        ("Diesel",      "Ind_Diesel"),
        ("PNG",         "Ind_PNG"),
    ])

    # Energy Industries (on-site generation)
    ng_tj  = float(d.get("ng_tj", 0) or 0)
    coal_tj= float(d.get("coal_tj", 0) or 0)
    msw_pw = float(d.get("msw_pw", 0) or 0)    # MW installed for MSW
    # Coal: tCO2e = coal_tj * (CO2+CH4+N2O)
    egen_coal = coal_tj * (EF["EGen_Coal"]["co2"] + EF["EGen_Coal"]["ch4"] + EF["EGen_Coal"]["n2o"])
    egen_gas  = ng_tj   * (EF["EGen_NatGas"]["co2"] + EF["EGen_NatGas"]["ch4"] + EF["EGen_NatGas"]["n2o"])
    subs["Energy Generation"] = egen_coal + egen_gas

    return subs


# ─── TRANSPORT ────────────────────────────────────────────────────────────────
def calc_transport(d):
    """
    Mirrors: Base Year GHG Inventory → Transport Sector (Fuel Sales Approach)
    Conversion: fuel (kl or tonne) → TJ → tCO2e
    """
    subs = {}

    def road_emit():
        # Petrol (kl) → TJ
        pet_kl = float(d.get("t_pet", 0) or 0)
        pet_tj = pet_kl * FUEL_CONV["Petrol"]["kl_to_tj"]
        e_pet  = pet_tj * ef_total_transport("Trans_Petrol")

        # Diesel (kl) → TJ
        die_kl = float(d.get("t_die", 0) or 0)
        die_tj = die_kl * FUEL_CONV["Diesel"]["kl_to_tj"]
        e_die  = die_tj * ef_total_transport("Trans_Diesel")

        # CNG (tonne) → TJ
        cng_t  = float(d.get("t_cng", 0) or 0)
        cng_tj = cng_t  * FUEL_CONV["CNG"]["t_to_tj"]
        e_cng  = cng_tj * ef_total_transport("Trans_CNG")

        # Auto LPG (tonne) → TJ
        alpg_t = float(d.get("t_alpg", 0) or 0)
        alpg_tj= alpg_t * FUEL_CONV["AutoLPG"]["t_to_tj"]
        e_alpg = alpg_tj * ef_total_transport("Trans_AutoLPG")

        # Electricity (MWh)
        elec_mwh = float(d.get("t_elec", 0) or 0)
        e_elec   = elec_mwh * EF["Trans_Electricity"]["co2"]

        return e_pet + e_die + e_cng + e_alpg + e_elec

    def rail_emit():
        die_kl  = float(d.get("r_die", 0) or 0)
        die_tj  = die_kl * FUEL_CONV["Diesel"]["kl_to_tj"]
        e_die   = die_tj * ef_total_transport("Trans_Diesel")
        elec_mwh= float(d.get("r_elec", 0) or 0)
        e_elec  = elec_mwh * EF["Trans_Electricity"]["co2"]
        return e_die + e_elec

    def water_emit():
        pet_kl = float(d.get("w_pet", 0) or 0)
        die_kl = float(d.get("w_die", 0) or 0)
        e = (pet_kl * FUEL_CONV["Petrol"]["kl_to_tj"] * ef_total_transport("Trans_Petrol") +
             die_kl * FUEL_CONV["Diesel"]["kl_to_tj"] * ef_total_transport("Trans_Diesel"))
        return e

    def aviation_emit():
        av_gas_kl= float(d.get("av_gas", 0) or 0)
        av_jet_kl= float(d.get("av_jet", 0) or 0)
        e = (av_gas_kl * FUEL_CONV["AvGasoline"]["kl_to_tj"] * ef_total_transport("Trans_AvGasoline") +
             av_jet_kl * FUEL_CONV["JetKerosene"]["kl_to_tj"] * ef_total_transport("Trans_JetKerosene"))
        return e

    subs["On Road"]              = road_emit()
    subs["Railway"]              = rail_emit()
    subs["Water Borne Navigation"]= water_emit()
    subs["Aviation"]             = aviation_emit()
    return subs


def ef_total_transport(key):
    e = EF[key]
    if e["unit"] == "t/MWh":
        return e["co2"]
    return e["co2"] + e["ch4"] + e["n2o"]


# ─── SOLID WASTE ──────────────────────────────────────────────────────────────
def calc_solid_waste(d):
    """
    Mirrors: E. Solid Waste → Landfill CH4 (IPCC FOD method)
    CH4_landfill = MSW_landfill × DOC_weighted × DOCF × F × 16/12 × (1-OX) × MCF
    Biogas/Composting CH4 from organic fraction
    """
    sw_tot = float(d.get("sw_tot", 0) or 0)   # tonne/day total MSW
    if sw_tot <= 0:
        return {"Solid Waste Disposal": 0.0, "Organic Waste Treatment": 0.0}

    sw_tpa = sw_tot * 365  # tonne/year

    # Waste fractions (default from E. Solid Waste sheet)
    f_food  = float(d.get("sw_food_frac",  0.726) or 0.726)
    f_paper = float(d.get("sw_paper_frac", 0.035) or 0.035)
    f_other = 1.0 - f_food - f_paper

    # Landfill fraction
    lfm = float(d.get("sw_lfm", 0.85) or 0.85)    # managed landfill %
    lfu = float(d.get("sw_lfu", 0.0)  or 0.0)      # unmanaged %
    sw_landfill_tpa = sw_tpa * (lfm + lfu)

    # Weighted DOC
    doc_w = (f_food * SW_DOC["food"] + f_paper * SW_DOC["paper"] +
             f_other * SW_DOC.get("rubber", 0.0))
    doc_w = max(doc_w, 0.05)

    # IPCC FOD constants (from E. Solid Waste sheet R89-94)
    docf = 0.6    # fraction of DOC ultimately decomposed
    f_ch4 = 0.5   # fraction of CH4 in landfill gas
    ox = 0.1      # oxidation factor
    mcf_managed = 1.0  # managed landfill

    # CH4 emissions (tCH4)
    ch4_gen = sw_landfill_tpa * doc_w * docf * f_ch4 * (16.0/12.0) * mcf_managed
    ch4_emit = ch4_gen * (1.0 - ox)  # subtract oxidised

    # Collection / recovery (default 0 unless input)
    collection_eff = float(d.get("sw_gas_collection", 0) or 0)
    ch4_recovered  = ch4_emit * collection_eff
    ch4_net = ch4_emit - ch4_recovered

    tco2e_landfill = ch4_net * GWP_CH4

    # Organic waste treatment (incineration / composting)
    inc_frac = float(d.get("sw_inc", 0.004) or 0.004)
    inc_tpa  = sw_tpa * inc_frac
    # Incineration: IPCC default EF 91.7 tCO2/TJ for non-biomass MSW
    # Using simplified: 0.5 tCO2e/t for non-biomass fraction
    tco2e_inc = inc_tpa * 0.5

    return {
        "Solid Waste Disposal":    tco2e_landfill,
        "Organic Waste Treatment": tco2e_inc,
    }


# ─── WASTEWATER ───────────────────────────────────────────────────────────────
def calc_wastewater(d):
    """
    Mirrors: F. Waste water Emission
    CH4 = BOD × B0 × MCF_weighted
    N2O from effluent discharge (protein-based)
    """
    population = float(d.get("population", 0) or 0)
    lpcd       = float(d.get("ww_lpcd", 135) or 135)   # L/person/day
    bod_pc     = float(d.get("ww_bod", 34) or 34)       # g BOD/person/day (sheet default 34)
    tn_pc      = float(d.get("ww_tn", 0.026) or 0.026)  # kg N/person/day (Table default)

    # BOD total (kg/year)
    bod_total = population * bod_pc / 1000.0 * 365  # kg/year
    # Total Nitrogen
    tn_total  = population * tn_pc * 365             # kg/year
    # Industrial co-discharge factor (from sheet: 1.056521739)
    co_factor = float(d.get("ww_co_factor", 1.0565) or 1.0565)
    bod_total *= co_factor
    tn_total  *= co_factor

    # Treatment fractions
    f_aer_not_well  = float(d.get("ww_aer", 0.17) or 0.17)
    f_anaerobic_r   = float(d.get("ww_uasb", 0.09) or 0.09)
    f_septic        = float(d.get("ww_sep", 0.0)  or 0.0)
    f_open          = float(d.get("ww_open", 0.06) or 0.06)
    f_untreated     = max(0.0, 1.0 - f_aer_not_well - f_anaerobic_r - f_septic - f_open)

    # B0 = 0.6 kg CH4/kg BOD (max capacity, from sheet R62)
    b0 = 0.6

    # CH4 by treatment type (kg CH4/yr)
    ch4_aer = bod_total * f_aer_not_well * b0 * WW_MCF["aerobic_ponds"]      # 0.0 (well managed)
    ch4_aer_nw = bod_total * f_aer_not_well * b0 * 0.3   # not well managed MCF=0.3
    ch4_anaer = bod_total * f_anaerobic_r * b0 * WW_MCF["anaerobic_reactor"]
    ch4_septic= bod_total * f_septic * b0 * WW_MCF["septic"]
    ch4_open  = bod_total * f_open   * b0 * WW_MCF["open_discharge"]
    ch4_unt   = bod_total * f_untreated * b0 * 0.1

    ch4_total_kg = ch4_aer_nw + ch4_anaer + ch4_septic + ch4_open + ch4_unt
    ch4_total_t  = ch4_total_kg / 1000.0
    tco2e_ch4    = ch4_total_t * GWP_CH4

    # N2O from effluent (IPCC 2019 Eq 6.9)
    # N2O-N from effluent = TN * EF_effluent (0.005 kg N2O-N/kg N)
    n2o_n = tn_total * 0.005   # kg N2O-N/yr
    n2o_t = n2o_n * (44.0/28.0) / 1000.0  # tN2O/yr
    tco2e_n2o = n2o_t * GWP_N2O

    return {"Waste water": tco2e_ch4 + tco2e_n2o}


# ─── AFOLU ────────────────────────────────────────────────────────────────────
def calc_afolu(d):
    """
    Mirrors: H. AFOLU sheet
    Enteric fermentation + Manure management CH4/N2O
    Land use change (simplified)
    """
    total_enteric = 0.0
    total_manure  = 0.0

    livestock_map = {
        "dairy_cow":   ("af_dc",  "dairy_cow_indigenous"),
        "nondairy_cow":("af_ndc", "nondairy_cow_adult"),
        "buffalo_d":   ("af_bufd","dairy_buffalo"),
        "buffalo_nd":  ("af_bufnd","dairy_cow_indigenous"),
        "sheep":       ("af_sheep","sheep"),
        "goat":        ("af_goat","goat"),
    }
    for ltype, (form_key, ef_key) in livestock_map.items():
        heads = float(d.get(form_key, 0) or 0)
        if heads > 0:
            # Enteric CH4
            ef_e = AFOLU_ENTERIC.get(ef_key, 0)
            ch4_e_t = heads * ef_e / 1000.0  # kg/head/yr → t/yr
            total_enteric += ch4_e_t * GWP_CH4
            # Manure CH4
            ef_m = AFOLU_MANURE_CH4.get(ef_key, 0)
            ch4_m_t = heads * ef_m / 1000.0
            total_manure += ch4_m_t * GWP_CH4

    # Wetland rice CH4 (simple: area × EF)
    wet_ha = float(d.get("af_wet", 0) or 0)
    # IPCC default EF: 1.3 kgCH4/ha/day, 120 day season
    ch4_wet_t = wet_ha * 1.3 * 120 / 1000.0 if wet_ha > 0 else 0.0
    total_wetland = ch4_wet_t * GWP_CH4

    # Forestland CO2 sequestration (negative emissions)
    forest_ha = float(d.get("af_fd", 0) or 0)
    # Average growth 5 tCO2/ha/yr for tropical moist (IPCC Tier 1)
    seq_forest = -forest_ha * 5.0 if forest_ha > 0 else 0.0

    # Grassland / managed land (simplified)
    grass_ha  = float(d.get("af_fm", 0) or 0)
    other_ha  = float(d.get("af_fo", 0) or 0)
    seq_land   = -(grass_ha + other_ha) * 1.5 if (grass_ha + other_ha) > 0 else 0.0

    return {
        "Live Stock":          total_enteric + total_manure + total_wetland,
        "Land Management":     seq_forest + seq_land,
        "Aggregate Sources":   0.0,
    }


# ─── IPPU ─────────────────────────────────────────────────────────────────────
def calc_ippu(d):
    """
    Mirrors: G. IPPU sheet
    Mineral + Chemical + Metal industries
    """
    subs = {}

    # Cement (clinker-based, IPCC Eq 2.2 Tier 2)
    clinker_t = float(d.get("ip_clink", 0) or 0)
    cfrac     = float(d.get("ip_cfrac", 1.0) or 1.0)   # clinker-to-cement ratio default 1
    subs["Mineral Industry"] = clinker_t * IPPU_EF["cement_clinker"] + \
                                float(d.get("ip_lime", 0) or 0) * IPPU_EF["lime_high_ca"] + \
                                float(d.get("ip_ls", 0) or 0)   * IPPU_EF["limestone"]

    # Chemical (ammonia, HNO3)
    nh3_t   = float(d.get("ip_nh3",  0) or 0)
    hno3_t  = float(d.get("ip_hno3", 0) or 0)
    n2o_hno3 = hno3_t * IPPU_EF["hno3_n2o"] / 1000.0 * GWP_N2O   # tN2O → tCO2e
    subs["Chemical Industry"] = nh3_t * IPPU_EF["ammonia"] + n2o_hno3

    # Metal (steel BOF/EAF)
    bof_t = float(d.get("ip_bof", 0) or 0)
    eaf_t = float(d.get("ip_eaf", 0) or 0)
    subs["Metal Industry"] = bof_t * IPPU_EF["steel_bof"] + eaf_t * IPPU_EF["steel_eaf"]

    subs["Non-Energy Products"]             = 0.0
    subs["Ozone Depleting Substances"]      = 0.0
    subs["Other Product Manufacture and Use"]= 0.0

    return subs


# ─── BAU PROJECTIONS ─────────────────────────────────────────────────────────
def calc_bau(base_emissions_by_sector, d, year):
    """
    Mirrors: BAU Scenario sheet
    Projection = Base * growth_factor^(year-base_year)
    Growth method: Population Growth Rate (as selected in BAU Scenario sheet)
    Special case: Railway stays flat from 2040 onwards (as in sheet R19)
    """
    base_year   = int(d.get("base_year", 2025))
    growth_rate = float(d.get("growth_rate", 0.03) or 0.03)
    n = year - base_year
    factor = (1 + growth_rate) ** n

    projected = {}
    for sector, subsectors in base_emissions_by_sector.items():
        projected[sector] = {}
        for sub, val in subsectors.items():
            if sector == "Transport" and sub == "Railway" and year >= 2040:
                # Railway flat from 2040 (as seen in BAU Scenario R19)
                factor_sub = (1 + growth_rate) ** (2040 - base_year)
            else:
                factor_sub = factor
            projected[sector][sub] = val * factor_sub
    return projected


# ─── TARGET SETTING ──────────────────────────────────────────────────────────
def calc_targets(bau_totals, d):
    """
    Mirrors: Target Setting sheet
    target_val = BAU - (target_pct * BAU_base)
    """
    base_year   = int(d.get("base_year", 2025))
    target_year = int(d.get("target_year", 2050))
    interim1    = int(d.get("interim1", 2030))
    interim2    = int(d.get("interim2", 2040))
    target_pct  = float(d.get("target_pct", 0.45) or 0.45)

    base_total = bau_totals.get(base_year, 0)

    targets = {}
    for yr, bau in bau_totals.items():
        if yr == base_year:
            targets[yr] = base_total
        else:
            # Linear % reduction scaling to target_year
            frac = min(1.0, (yr - base_year) / max(1, (target_year - base_year)))
            reduction = base_total * target_pct * frac
            targets[yr] = bau - reduction
    return targets


# ─── EMISSION REDUCTION GRAPH — E&P and High Ambition ────────────────────────
def calc_scenarios(base_emissions_by_sector, bau_by_year, d):
    """
    Mirrors: Emission Reduction- Graph sheet
    For each sector×subsector×year:
      EP_reduction  = BAU × ep_pct  (user slider)
      HA_reduction  = BAU × ha_pct  (user slider)
      EP_emission   = BAU - EP_reduction
      HA_emission   = BAU - HA_reduction
    Then aggregate.
    """
    years = sorted(bau_by_year.keys())

    ep_total = {}
    ha_total = {}

    SUBSECTOR_KEYS = {
        "Residential":            "Residential",
        "Commercial":             "Commercial",
        "Public & Institutional": "Public___Institutional",
        "Industrial":             "Industrial",
        "Energy Generation":      "Energy_Generation",
        "On Road":                "On_Road",
        "Railway":                "Railway",
        "Water Borne Navigation": "Water_Borne_Navigation",
        "Aviation":               "Aviation",
        "Solid Waste Disposal":   "Solid_Waste_Disposal",
        "Organic Waste Treatment":"Organic_Waste_Treatment",
        "Waste water":            "Waste_water",
        "Live Stock":             "Live_Stock",
        "Land Management":        "Land_Management",
        "Aggregate Sources":      "Aggregate_Sources",
        "Mineral Industry":       "Mineral_Industry",
        "Chemical Industry":      "Chemical_Industry",
        "Metal Industry":         "Metal_Industry",
    }

    for yr in years:
        bau_yr = bau_by_year[yr]
        ep_yr  = 0.0
        ha_yr  = 0.0
        for sector, subsectors in bau_yr.items():
            for sub, bau_val in subsectors.items():
                key = SUBSECTOR_KEYS.get(sub, sub.replace(" ","_").replace("&","").replace("/","_"))
                # Get slider values (default from Excel sheet E&P/HA defaults)
                ep_pct_key = f"ep_pct_{key}"
                ha_pct_key = f"ha_pct_{key}"
                ep_pct = float(d.get(ep_pct_key, _ep_default(sub)) or _ep_default(sub)) / 100.0
                ha_pct = float(d.get(ha_pct_key, _ha_default(sub)) or _ha_default(sub)) / 100.0
                ep_yr += bau_val * (1.0 - ep_pct)
                ha_yr += bau_val * (1.0 - ha_pct)
        ep_total[yr] = ep_yr
        ha_total[yr] = ha_yr

    return ep_total, ha_total


def _ep_default(subsector):
    """Default E&P reduction % from Emission Reduction-Graph sheet (col E)."""
    EP_DEFAULTS = {
        "Residential": 10, "Commercial": 5, "Public & Institutional": 5,
        "Industrial": 5, "Energy Generation": 0, "On Road": 5, "Railway": 5,
        "Water Borne Navigation": 5, "Aviation": 10,
        "Solid Waste Disposal": 5, "Organic Waste Treatment": 5,
        "Waste water": 5, "Live Stock": 5, "Land Management": 5,
        "Mineral Industry": 5, "Chemical Industry": 5, "Metal Industry": 5,
    }
    return EP_DEFAULTS.get(subsector, 5)


def _ha_default(subsector):
    """Default High Ambition reduction % from Emission Reduction-Graph sheet (col H)."""
    HA_DEFAULTS = {
        "Residential": 30, "Commercial": 30, "Public & Institutional": 30,
        "Industrial": 30, "Energy Generation": 20, "On Road": 35, "Railway": 20,
        "Water Borne Navigation": 20, "Aviation": 30,
        "Solid Waste Disposal": 40, "Organic Waste Treatment": 40,
        "Waste water": 40, "Live Stock": 15, "Land Management": 15,
        "Mineral Industry": 20, "Chemical Industry": 20, "Metal Industry": 20,
    }
    return HA_DEFAULTS.get(subsector, 20)


# ─── MITIGATION BUDGET (Strategies & Cost) ────────────────────────────────────
def calc_mitigation_budget(base_by_sector, bau_by_year, ha_by_year, d):
    """
    Mirrors: Stratergies & Cost + Dashboard- Scenario Comparison
    For each sector: GHG reduced = BAU_targetyear - HA_targetyear
    Investment = GHG_reduced * cost_per_tonne (sector-specific)
    """
    target_year = int(d.get("target_year", 2050))
    bau_ty = bau_by_year.get(target_year, {})
    ha_ty  = ha_by_year.get(target_year, 0)

    # Sector-level BAU aggregates at target year
    sector_bau = {}
    for sector, subsectors in bau_ty.items():
        sector_bau[sector] = sum(subsectors.values())

    # HA total at target year
    bau_total_ty = sum(sector_bau.values())
    ha_total_ty  = ha_by_year.get(target_year, bau_total_ty)

    budget_rows = []
    total_reduced = 0.0
    total_inv = 0.0

    sector_map = {
        "Energy Sector":  ("Buildings", ABATEMENT_COST["Buildings"]),
        "Transport":      ("Transport", ABATEMENT_COST["Transport"]),
        "Waste":          ("Waste",     ABATEMENT_COST["Waste"]),
        "Wastewater":     ("Wastewater",ABATEMENT_COST["Wastewater"]),
        "AFOLU":          ("AFOLU",     ABATEMENT_COST["AFOLU"]),
        "IPPU":           ("IPPU",      ABATEMENT_COST["IPPU"]),
    }

    for sector, bau_val in sector_bau.items():
        display, cost_t = sector_map.get(sector, (sector, 2000))
        # Proportional HA reduction
        if bau_total_ty > 0:
            ha_val = ha_total_ty * (bau_val / bau_total_ty)
        else:
            ha_val = 0.0

        # User-defined reduction %
        red_pct_key = f"ha_pct_{display.replace(' ','_')}"
        red_pct = float(d.get(red_pct_key, 20) or 20) / 100.0
        reduced = bau_val * red_pct
        inv = reduced * cost_t / 1e7  # ₹ Crore

        total_reduced += reduced
        total_inv     += inv
        budget_rows.append({
            "Sector":               display,
            "BAU (t CO2e)":         round(bau_val),
            "Reduction %":          f"{red_pct*100:.0f}%",
            "GHG Reduced (t CO2e)": round(reduced),
            "Investment (Crore)":   f"₹{inv:,.1f}",
        })

    budget_rows.append({
        "Sector":               "TOTAL",
        "BAU (t CO2e)":         round(bau_total_ty),
        "Reduction %":          f"{(total_reduced/bau_total_ty*100):.1f}%" if bau_total_ty else "—",
        "GHG Reduced (t CO2e)": round(total_reduced),
        "Investment (Crore)":   f"₹{total_inv:,.1f}",
    })

    return budget_rows, total_inv


# ─── MILESTONE TABLE (Target Setting) ─────────────────────────────────────────
def calc_milestones(bau_totals, ep_totals, ha_totals, targets, d):
    """
    Mirrors: Target Setting + Emission Reduction-Graph R99-103
    """
    rows = []
    base_year = int(d.get("base_year", 2025))
    years = sorted(bau_totals.keys())
    for yr in years:
        bau = bau_totals[yr]
        tgt = targets.get(yr, bau)
        ha  = ha_totals.get(yr, bau)
        ep  = ep_totals.get(yr, bau)
        base= bau_totals[base_year]
        req_pct = (bau - tgt) / bau * 100 if bau > 0 else 0
        ach_pct = (bau - ha)  / bau * 100 if bau > 0 else 0
        status  = "On Track" if ha <= tgt else "Gap"
        rows.append({
            "year":         yr,
            "bau":          round(bau / 1e6, 2),
            "target":       round(tgt / 1e6, 2),
            "ep":           round(ep  / 1e6, 2),
            "ha":           round(ha  / 1e6, 2),
            "required_pct": f"{req_pct:.1f}%",
            "achieved_pct": f"{ach_pct:.1f}%",
            "status":       status,
        })
    return rows


# ─── PLOTLY CHARTS ────────────────────────────────────────────────────────────
NAVY  = "#1a2744"
TEAL  = "#00b4a6"
AMBER = "#f59e0b"
GREEN = "#10b981"
RED   = "#ef4444"
BLUE  = "#3b82f6"
CHART_FONT = dict(family="DM Sans, sans-serif", size=12, color="#444")


def make_trajectory_chart(bau_totals, ep_totals, ha_totals, targets, years):
    """Mirrors Emission Reduction-Graph dashboard chart (rows 43-47, 99-103)."""
    fig = go.Figure()

    ys = sorted(years)
    scale = 1e6  # tCO2e → MtCO2e

    fig.add_trace(go.Scatter(
        x=ys, y=[bau_totals[y]/scale for y in ys],
        name="Reference (BAU)", mode="lines+markers",
        line=dict(color=NAVY, width=2.5, dash="solid"),
        marker=dict(size=6)
    ))
    fig.add_trace(go.Scatter(
        x=ys, y=[ep_totals.get(y, bau_totals[y])/scale for y in ys],
        name="Existing & Planned (E&P)", mode="lines+markers",
        line=dict(color=BLUE, width=2, dash="dot"),
        marker=dict(size=6)
    ))
    fig.add_trace(go.Scatter(
        x=ys, y=[ha_totals.get(y, bau_totals[y])/scale for y in ys],
        name="High Ambition", mode="lines+markers",
        line=dict(color=TEAL, width=2.5),
        marker=dict(size=7)
    ))
    fig.add_trace(go.Scatter(
        x=ys, y=[targets.get(y, bau_totals[y])/scale for y in ys],
        name="Target Pathway", mode="lines+markers",
        line=dict(color=GREEN, width=2, dash="dash"),
        marker=dict(symbol="diamond", size=8)
    ))

    # Shaded gap area between BAU and HA
    ys_r = list(reversed(ys))
    fig.add_trace(go.Scatter(
        x=ys + ys_r,
        y=[bau_totals[y]/scale for y in ys] + [ha_totals.get(y, bau_totals[y])/scale for y in ys_r],
        fill="toself", fillcolor="rgba(0,180,166,0.10)",
        line=dict(color="rgba(0,0,0,0)"), showlegend=False, hoverinfo="skip"
    ))

    fig.update_layout(
        title=dict(text="GHG Emission Trajectory (Mt CO₂e)", font=dict(size=14)),
        xaxis=dict(title="Year", gridcolor="#f0f0f0"),
        yaxis=dict(title="Mt CO₂e", gridcolor="#f0f0f0"),
        legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1),
        plot_bgcolor="white", paper_bgcolor="white",
        font=CHART_FONT, margin=dict(l=50, r=20, t=60, b=40),
        hovermode="x unified"
    )
    return json.loads(fig.to_json())


def make_pie_chart(base_by_sector):
    """Mirrors Dashboard- BAU-City sector breakdown pie chart."""
    labels, values = [], []
    sector_colors = {
        "Energy Sector":  NAVY,
        "Transport":      TEAL,
        "Waste":          AMBER,
        "Wastewater":     BLUE,
        "AFOLU":          GREEN,
        "IPPU":           "#8b5cf6",
    }
    for sector, subs in base_by_sector.items():
        total = sum(subs.values())
        if total > 0:
            labels.append(sector)
            values.append(total)

    fig = go.Figure(go.Pie(
        labels=labels, values=values,
        hole=0.42,
        marker=dict(colors=[sector_colors.get(l, "#aaa") for l in labels]),
        textinfo="label+percent",
        textfont=dict(size=11),
        hovertemplate="%{label}<br>%{value:,.0f} tCO₂e<br>%{percent}<extra></extra>"
    ))
    fig.update_layout(
        title=dict(text="Base Year Emissions by Sector", font=dict(size=14)),
        showlegend=False, font=CHART_FONT,
        plot_bgcolor="white", paper_bgcolor="white",
        margin=dict(l=10, r=10, t=50, b=10)
    )
    return json.loads(fig.to_json())


def make_bar_chart(bau_totals, ep_totals, ha_totals, targets):
    """Grouped bar chart — mirrors Dashboard- Scenario Comparison."""
    years = sorted(bau_totals.keys())
    scale = 1e6

    fig = go.Figure()
    fig.add_trace(go.Bar(name="BAU",          x=years,
                         y=[bau_totals[y]/scale for y in years], marker_color=NAVY))
    fig.add_trace(go.Bar(name="E&P",          x=years,
                         y=[ep_totals.get(y,0)/scale for y in years], marker_color=BLUE))
    fig.add_trace(go.Bar(name="High Ambition",x=years,
                         y=[ha_totals.get(y,0)/scale for y in years], marker_color=TEAL))
    fig.add_trace(go.Bar(name="Target",       x=years,
                         y=[targets.get(y,0)/scale for y in years],   marker_color=GREEN))

    fig.update_layout(
        barmode="group",
        title=dict(text="Scenario Comparison by Year", font=dict(size=14)),
        xaxis=dict(title="Year", type="category"),
        yaxis=dict(title="Mt CO₂e", gridcolor="#f0f0f0"),
        legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1),
        plot_bgcolor="white", paper_bgcolor="white",
        font=CHART_FONT, margin=dict(l=50, r=20, t=60, b=40),
    )
    return json.loads(fig.to_json())


def make_budget_chart(budget_rows):
    """Mitigation cost waterfall by sector."""
    rows = [r for r in budget_rows if r["Sector"] != "TOTAL"]
    sectors = [r["Sector"] for r in rows]
    reductions = [r["GHG Reduced (t CO2e)"] / 1e6 for r in rows]

    fig = go.Figure(go.Bar(
        x=sectors, y=reductions,
        marker_color=TEAL,
        text=[f"{v:.2f} Mt" for v in reductions],
        textposition="outside",
        hovertemplate="%{x}<br>Reduced: %{y:.3f} Mt CO₂e<extra></extra>"
    ))
    fig.update_layout(
        title=dict(text="GHG Reduction Potential by Sector (Target Year)", font=dict(size=14)),
        xaxis=dict(title="Sector"),
        yaxis=dict(title="Mt CO₂e Reduced", gridcolor="#f0f0f0"),
        plot_bgcolor="white", paper_bgcolor="white",
        font=CHART_FONT, margin=dict(l=50, r=20, t=60, b=80),
    )
    return json.loads(fig.to_json())


def make_subsector_bar(base_by_sector):
    """Stacked sub-sector bar."""
    sectors, totals = [], []
    for sector, subs in base_by_sector.items():
        t = sum(subs.values())
        if t > 0:
            sectors.append(sector)
            totals.append(t / 1e6)

    fig = go.Figure(go.Bar(
        x=sectors, y=totals,
        marker_color=[NAVY, TEAL, AMBER, BLUE, GREEN, "#8b5cf6"][:len(sectors)],
        text=[f"{v:.2f}" for v in totals],
        textposition="outside",
    ))
    fig.update_layout(
        title=dict(text="Base Year Emissions by Sector (Mt CO₂e)", font=dict(size=14)),
        xaxis=dict(title="Sector"),
        yaxis=dict(title="Mt CO₂e", gridcolor="#f0f0f0"),
        plot_bgcolor="white", paper_bgcolor="white",
        font=CHART_FONT, margin=dict(l=50, r=20, t=60, b=80),
    )
    return json.loads(fig.to_json())


# ═══════════════════════════════════════════════════════════════════════════════
# FLASK ROUTES
# ═══════════════════════════════════════════════════════════════════════════════

@app.route("/")
def index():
    return render_template("index.html", cities=INDIA_CITIES)


@app.route("/results")
def results():
    return render_template("results.html")


@app.route("/api/cities")
def api_cities():
    return jsonify(INDIA_CITIES)


@app.route("/api/calculate", methods=["POST"])
def api_calculate():
    d = request.get_json(force=True)

    # ── Step 1: Base Year Emissions ──────────────────────────────────────────
    bldg   = calc_buildings(d)
    trans  = calc_transport(d)
    sw     = calc_solid_waste(d)
    ww     = calc_wastewater(d)
    afolu  = calc_afolu(d)
    ippu   = calc_ippu(d)

    base_by_sector = {
        "Energy Sector": bldg,
        "Transport":     trans,
        "Waste":         sw,
        "Wastewater":    ww,
        "AFOLU":         afolu,
        "IPPU":          ippu,
    }

    base_total = sum(sum(s.values()) for s in base_by_sector.values())

    # ── Step 2: BAU Projections ──────────────────────────────────────────────
    base_year   = int(d.get("base_year",   2025))
    interim1    = int(d.get("interim1",    2030))
    interim2    = int(d.get("interim2",    2040))
    target_year = int(d.get("target_year", 2050))
    years = sorted(set([base_year, interim1, interim2, target_year]))

    bau_by_year = {}
    for yr in years:
        bau_by_year[yr] = calc_bau(base_by_sector, d, yr)

    bau_totals = {yr: sum(sum(s.values()) for s in bau_by_year[yr].values())
                  for yr in years}

    # ── Step 3: Target Setting ───────────────────────────────────────────────
    targets = calc_targets(bau_totals, d)

    # ── Step 4: Scenario Emissions (E&P and High Ambition) ──────────────────
    ep_totals, ha_totals = calc_scenarios(base_by_sector, bau_by_year, d)

    # ── Step 5: Mitigation Budget ────────────────────────────────────────────
    budget_rows, total_inv = calc_mitigation_budget(
        base_by_sector, bau_by_year, ha_totals, d)

    # ── Step 6: Milestones ───────────────────────────────────────────────────
    milestones = calc_milestones(bau_totals, ep_totals, ha_totals, targets, d)

    # ── Step 7: KPIs ─────────────────────────────────────────────────────────
    population = max(float(d.get("population", 1) or 1), 1)
    area_sqkm  = max(float(d.get("area_sqkm",  1) or 1), 1)
    kpis = {
        "base_total_mt": round(base_total / 1e6, 2),
        "per_capita":    round(base_total / population, 2),
        "per_sqkm":      round(base_total / area_sqkm / 1000, 2),
        "bau_end_mt":    round(bau_totals.get(target_year, 0) / 1e6, 2),
        "ep_end_mt":     round(ep_totals.get(target_year, 0) / 1e6, 2),
        "ha_end_mt":     round(ha_totals.get(target_year, 0) / 1e6, 2),
        "target_mt":     round(targets.get(target_year, 0) / 1e6, 2),
        "total_inv":     round(total_inv, 1),
        "base_year":     base_year,
        "target_year":   target_year,
    }

    # ── Step 8: Sector Detail Table ──────────────────────────────────────────
    sector_detail = []
    for sector, subs in base_by_sector.items():
        for sub, val in subs.items():
            if val != 0:
                sector_detail.append({
                    "sector":    f"{sector} – {sub}",
                    "emissions": round(val),
                    "share":     f"{val/base_total*100:.1f}%" if base_total > 0 else "0%"
                })
    sector_detail.sort(key=lambda x: -x["emissions"])

    # ── Step 9: Charts ───────────────────────────────────────────────────────
    charts = {
        "trajectory": make_trajectory_chart(bau_totals, ep_totals, ha_totals, targets, years),
        "pie":        make_pie_chart(base_by_sector),
        "bar_group":  make_bar_chart(bau_totals, ep_totals, ha_totals, targets),
        "budget":     make_budget_chart(budget_rows),
        "subsector":  make_subsector_bar(base_by_sector),
    }

    return jsonify({
        "kpis":          kpis,
        "charts":        charts,
        "milestones":    milestones,
        "sector_detail": sector_detail,
        "budget":        budget_rows,
    })


@app.route("/api/download/excel", methods=["POST"])
def download_excel():
    """
    Generate styled Excel export mirroring the structure of the ASCENT workbook.
    4 sheets: Summary, Base Year Emissions, BAU Projections, Mitigation Budget
    """
    d = request.get_json(force=True)

    # Re-run calculations
    bldg  = calc_buildings(d)
    trans = calc_transport(d)
    sw    = calc_solid_waste(d)
    ww    = calc_wastewater(d)
    afolu = calc_afolu(d)
    ippu  = calc_ippu(d)

    base_by_sector = {
        "Energy Sector": bldg, "Transport": trans,
        "Waste": sw, "Wastewater": ww, "AFOLU": afolu, "IPPU": ippu,
    }
    base_total = sum(sum(s.values()) for s in base_by_sector.values())
    population  = max(float(d.get("population", 1) or 1), 1)
    area_sqkm   = max(float(d.get("area_sqkm", 1) or 1), 1)

    base_year   = int(d.get("base_year",   2025))
    interim1    = int(d.get("interim1",    2030))
    interim2    = int(d.get("interim2",    2040))
    target_year = int(d.get("target_year", 2050))
    years = sorted(set([base_year, interim1, interim2, target_year]))

    bau_by_year = {yr: calc_bau(base_by_sector, d, yr) for yr in years}
    bau_totals  = {yr: sum(sum(s.values()) for s in bau_by_year[yr].values()) for yr in years}
    targets     = calc_targets(bau_totals, d)
    ep_totals, ha_totals = calc_scenarios(base_by_sector, bau_by_year, d)
    budget_rows, total_inv = calc_mitigation_budget(base_by_sector, bau_by_year, ha_totals, d)

    wb = Workbook()

    HDR_FILL   = PatternFill("solid", fgColor="1a2744")
    TEAL_FILL  = PatternFill("solid", fgColor="00b4a6")
    LIGHT_FILL = PatternFill("solid", fgColor="e8f4f3")
    HDR_FONT   = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    BOLD       = Font(name="Calibri", bold=True, size=10)
    NORMAL     = Font(name="Calibri", size=10)
    CENTER     = Alignment(horizontal="center", vertical="center")
    LEFT       = Alignment(horizontal="left",   vertical="center")
    thin = Side(style="thin", color="CCCCCC")
    thin_border= Border(left=thin, right=thin, top=thin, bottom=thin)

    city  = d.get("city", d.get("district","City"))
    state = d.get("state","")
    tier  = d.get("tier","District")

    def hdr_row(ws, row, cols, values):
        for c, v in zip(cols, values):
            cell = ws.cell(row=row, column=c, value=v)
            cell.fill = HDR_FILL; cell.font = HDR_FONT
            cell.alignment = CENTER; cell.border = thin_border

    def data_row(ws, row, cols, values, alt=False):
        fill = LIGHT_FILL if alt else PatternFill()
        for c, v in zip(cols, values):
            cell = ws.cell(row=row, column=c, value=v)
            if alt: cell.fill = fill
            cell.font = NORMAL; cell.alignment = LEFT; cell.border = thin_border

    # ── Sheet 1: Summary ───────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Summary"
    ws1.column_dimensions["A"].width = 30
    ws1.column_dimensions["B"].width = 22

    ws1.merge_cells("A1:B1")
    c = ws1["A1"]
    c.value = f"ASCENT GHG Inventory — {city}, {state}"
    c.font = Font(name="Calibri", bold=True, size=14, color="1a2744")
    c.alignment = CENTER
    ws1.row_dimensions[1].height = 28

    meta = [
        ("City / District", city), ("State", state), ("Government Tier", tier),
        ("Population", f"{int(population):,}"),
        ("Area (km²)", area_sqkm),
        ("Base Year", base_year), ("Target Year", target_year),
        ("Climate Zone", d.get("climate", "—")),
        ("Growth Rate", f"{float(d.get('growth_rate',0.03))*100:.1f}%"),
        ("Target Reduction", f"{float(d.get('target_pct',0.45))*100:.0f}%"),
        ("", ""),
        ("Base Year Total Emissions (tCO₂e)", round(base_total)),
        ("Per Capita (tCO₂e/person)", round(base_total/population, 2)),
        ("Per km² (tCO₂e/km²)", round(base_total/area_sqkm)),
        ("BAU at Target Year (Mt)", round(bau_totals.get(target_year,0)/1e6, 2)),
        ("E&P at Target Year (Mt)", round(ep_totals.get(target_year,0)/1e6, 2)),
        ("High Ambition at Target Year (Mt)", round(ha_totals.get(target_year,0)/1e6, 2)),
        ("GHG Reduction Needed (Mt)", round((bau_totals.get(target_year,0)-targets.get(target_year,0))/1e6, 2)),
        ("Total Mitigation Investment (₹ Cr)", round(total_inv, 1)),
    ]
    for r, (label, val) in enumerate(meta, start=3):
        ws1.cell(row=r, column=1, value=label).font = BOLD
        ws1.cell(row=r, column=2, value=val).font = NORMAL

    # ── Sheet 2: Base Year Emissions ──────────────────────────────────────
    ws2 = wb.create_sheet("Base Year Emissions")
    for col, w in zip("ABCD", [30, 28, 20, 15]):
        ws2.column_dimensions[chr(ord("A")+["ABCD".index(c) for c in "ABCD"][["ABCD".index(c) for c in "ABCD"].index(col.replace(col,col))])].width = w
    ws2.column_dimensions["A"].width = 28
    ws2.column_dimensions["B"].width = 28
    ws2.column_dimensions["C"].width = 20
    ws2.column_dimensions["D"].width = 14

    hdr_row(ws2, 1, [1,2,3,4],
            ["Sector","Sub-Sector","Emissions (tCO₂e)","Share (%)"])
    r = 2
    for sector, subs in base_by_sector.items():
        for sub, val in subs.items():
            share = val/base_total*100 if base_total > 0 else 0
            data_row(ws2, r, [1,2,3,4],
                     [sector, sub, round(val), f"{share:.1f}%"], alt=(r%2==0))
            r += 1
    data_row(ws2, r, [1,2,3,4],
             ["TOTAL", "", round(base_total), "100%"], alt=False)
    ws2.cell(row=r, column=1).font = BOLD
    ws2.cell(row=r, column=3).font = BOLD

    # ── Sheet 3: BAU & Scenarios ──────────────────────────────────────────
    ws3 = wb.create_sheet("BAU & Scenarios")
    ws3.column_dimensions["A"].width = 10
    for col in "BCDE": ws3.column_dimensions[col].width = 20

    hdr_row(ws3, 1, [1,2,3,4,5],
            ["Year","BAU (Mt CO₂e)","E&P (Mt CO₂e)","High Ambition (Mt)","Target (Mt)"])
    for r, yr in enumerate(years, start=2):
        data_row(ws3, r, [1,2,3,4,5], [
            yr,
            round(bau_totals[yr]/1e6, 3),
            round(ep_totals.get(yr,0)/1e6, 3),
            round(ha_totals.get(yr,0)/1e6, 3),
            round(targets.get(yr,0)/1e6, 3),
        ], alt=(r%2==0))

    # ── Sheet 4: Mitigation Budget ────────────────────────────────────────
    ws4 = wb.create_sheet("Mitigation Budget")
    ws4.column_dimensions["A"].width = 22
    ws4.column_dimensions["B"].width = 20
    ws4.column_dimensions["C"].width = 14
    ws4.column_dimensions["D"].width = 22
    ws4.column_dimensions["E"].width = 22

    hdr_row(ws4, 1, [1,2,3,4,5],
            ["Sector","BAU at Target Year (t)","Reduction %",
             "GHG Reduced (t CO₂e)","Investment (₹ Crore)"])
    for r, row in enumerate(budget_rows, start=2):
        data_row(ws4, r, [1,2,3,4,5], [
            row["Sector"],
            row["BAU (t CO2e)"],
            row["Reduction %"],
            row["GHG Reduced (t CO2e)"],
            row["Investment (Crore)"],
        ], alt=(r%2==0))
        if row["Sector"] == "TOTAL":
            for c in range(1,6):
                ws4.cell(row=r, column=c).font = BOLD
                ws4.cell(row=r, column=c).fill = TEAL_FILL

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"ASCENT_{city.replace(' ','_')}_{base_year}_{target_year}.xlsx"
    return send_file(buf, as_attachment=True, download_name=fname,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.route("/api/download/csv", methods=["POST"])
def download_csv():
    d = request.get_json(force=True)
    bldg  = calc_buildings(d); trans = calc_transport(d)
    sw    = calc_solid_waste(d); ww  = calc_wastewater(d)
    afolu = calc_afolu(d); ippu = calc_ippu(d)
    base_by_sector = {
        "Energy Sector": bldg, "Transport": trans,
        "Waste": sw, "Wastewater": ww, "AFOLU": afolu, "IPPU": ippu,
    }
    base_total = sum(sum(s.values()) for s in base_by_sector.values())
    base_year   = int(d.get("base_year",2025))
    interim1    = int(d.get("interim1",2030))
    interim2    = int(d.get("interim2",2040))
    target_year = int(d.get("target_year",2050))
    years = sorted(set([base_year, interim1, interim2, target_year]))
    bau_by_year = {yr: calc_bau(base_by_sector, d, yr) for yr in years}
    bau_totals  = {yr: sum(sum(s.values()) for s in bau_by_year[yr].values()) for yr in years}
    targets     = calc_targets(bau_totals, d)
    ep_totals, ha_totals = calc_scenarios(base_by_sector, bau_by_year, d)

    lines = ["Year,BAU (Mt),E&P (Mt),High Ambition (Mt),Target (Mt)"]
    for yr in years:
        lines.append(f"{yr},"
                     f"{bau_totals[yr]/1e6:.3f},"
                     f"{ep_totals.get(yr,0)/1e6:.3f},"
                     f"{ha_totals.get(yr,0)/1e6:.3f},"
                     f"{targets.get(yr,0)/1e6:.3f}")
    csv_text = "\n".join(lines)
    buf = io.BytesIO(csv_text.encode())
    buf.seek(0)
    city = d.get("city", d.get("district","city")).replace(" ","_")
    return send_file(buf, as_attachment=True,
                     download_name=f"ASCENT_{city}_scenarios.csv",
                     mimetype="text/csv")


if __name__ == "__main__":
    app.run(debug=True, host="0.0.0.0", port=5000)
